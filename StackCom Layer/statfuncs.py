import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from lxml import etree
import tkinter as tk
from tkinter import filedialog
import math
from tkinter import ttk

#the excel output file file path
file_path = os.path.join(os.getcwd(), 'Output.xlsx')

#the namespace used for the XDM documents
namespace = {'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd','a':'http://www.tresos.de/_projects/DataModel2/08/attribute.xsd'}

# Define expected headers for cleaning the Excel data
expected_trame_headers = {'FRAMES': ['Checked','Radical', 'Identifiant_T', 'Taille_Max_T', 'Mode_Transmission_T', 'Periode_T', 'UCE Emetteur', 'AEE10r3 Reseau_T']}

# Function to clean the Excel data and keep only the necessary columns
def cleanExcelFrameData(excel_file):
    df = pd.read_excel(excel_file, sheet_name='FRAMES', header=None, skiprows=2)  # Skip first two rows
    df.columns = df.iloc[0]  # Use the third row as column names
    df = df.iloc[1:]  
    
    headers = [col for col in df.columns if col in expected_trame_headers['FRAMES']]
    df = df[headers]
    
    # Filter rows based on 'checked' column
    df = df[df['Checked'] == 'X']
    
    return df

    # Function to get all the Excel data
def getFullFrameData(excel_file):
    df = pd.read_excel(excel_file, sheet_name='FRAMES', header=None, skiprows=2)  # Skip first two rows
    df.columns = df.iloc[0]  # Use the third row as column names
    df = df.iloc[1:]  
    headers = [col for col in df.columns if col in expected_trame_headers['FRAMES']]
    df = df[headers]
    return df

# Define expected headers for cleaning the Excel data
expected_signal_headers = {'SIGNALS': ['Checked','Radical_T', 'Position_octet_S', 'Position_bit_S', 'Taille_Max_S', 'Mnemonique_S', 'Valeur_Min_S', 'Valeur_Max_S', 'Resolution_S', 'Offset_S','PROD_INIT','CONS_INIT' ,'Emetteur','Nécessité de sécurisation par checksum et compteur de process']}

# Function to clean the Excel data and keep only the necessary columns
def cleanExcelSignalData(excel_file):
    df = pd.read_excel(excel_file, sheet_name='SIGNALS', header=None, skiprows=2)  
    df.columns = df.iloc[0]  
    df = df.iloc[1:]  
    headers = [col for col in df.columns if col in expected_signal_headers['SIGNALS']]
    df = df[headers]
    df = df[df['Checked'] == 'X']
    return df

# Function to get all the Excel data
def getFullSignalData(excel_file):
    df = pd.read_excel(excel_file, sheet_name='SIGNALS', header=None, skiprows=2)  
    df.columns = df.iloc[0]  
    df = df.iloc[1:]  
    headers = [col for col in df.columns if col in expected_signal_headers['SIGNALS']]
    df = df[headers]
    return df

#display the list of the frames in the interface on excel file selection
def display_frame_names(selected_excel_file,frame_entry):
    if selected_excel_file:
        cleaned_df = cleanExcelFrameData(selected_excel_file)
        frame_names = cleaned_df['Radical'].tolist()
        frame_entry['values'] = frame_names

#display the list of the signals in the interface on excel file selection
def display_signal_names(selected_excel_file,signal_entry):
    if selected_excel_file:
        cleaned_df = cleanExcelSignalData(selected_excel_file)
        full_names = cleaned_df.apply(lambda row: row['Mnemonique_S'] + "_" + row['Radical_T'], axis=1)
        signal_names = full_names.tolist()
        signal_entry['values'] = signal_names


#write data to the output excel file
def write_to_Excel(result_data, file_path, sheet_name):
    df = pd.DataFrame(result_data)

    if not os.path.exists(file_path):
        writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Apply formatting to header row 
        header_format = workbook.add_format({'bg_color': '5DADE2', 'align': 'center', 'valign': 'vcenter'})
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            column_width = max(len(str(value)) + 2, len(df[df.columns[col_num]].astype(str).max()) + 2)
            worksheet.set_column(col_num, col_num, column_width)
        
        # Apply formatting to the first column 
        first_column_format = workbook.add_format({'bg_color': '5DADE2', 'align': 'center', 'valign': 'vcenter'})
        worksheet.set_column(0, 0, 15)  # Set a specific width for the first column
        for row_num in range(1, df.shape[0] + 1):
            worksheet.write(row_num, 0, df.iloc[row_num - 1, 0], first_column_format)
            worksheet.set_row(row_num, len(str(df.iloc[row_num - 1, 0])) * 1.5)
        
        df.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=0, index=False, header=False)
        
        writer.save()
    else:
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book

        if sheet_name in pd.ExcelFile(file_path).sheet_names:
            sheet = book[sheet_name]
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=writer.sheets[sheet_name].max_row)

        else:
            # Create a new sheet if it doesn't exist
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        writer.save()

# Function responsible for emptying the designated Excel sheet
def clear_excel(sheet_name, completion_label):
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            sheet.delete_rows(2, sheet.max_row - 1)
        book.save(file_path)
        completion_label.config(text="Output File Cleared", fg="blue")  

    
#select the excel file from the interface
def browse_excel_frames(excel_file_entry,frame_entry):
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not excel_file_path:
        return
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(tk.END, excel_file_path)
    display_frame_names(excel_file_path,frame_entry)

#select the excel file from the interface
def browse_excel_signals(excel_file_entry,signal_entry):
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not excel_file_path:
        return
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(tk.END, excel_file_path)
    display_signal_names(excel_file_path,signal_entry)


#select the xdm file from the interface
def browse_xdm(xdm_file_entry):
    xdm_file_path = filedialog.askopenfilename(filetypes=[("XDM files", "*.xdm")])
    if not xdm_file_path:
        return
    xdm_file_entry.delete(0, tk.END)
    xdm_file_entry.insert(tk.END, xdm_file_path)

#execute functionality on button click
def verify_button_click(excel_file_path,xdm_file_path,frame_name):
    excel_file_path = excel_file_entry.get()
    xdm_file_path = xdm_file_entry.get()
    frame_name = frame_entry.get()

    verify_frame(excel_file_path, xdm_file_path, frame_name)
    completion_label.config(text="Output Created", fg="green")

#Ordering by index table CanIF
def ordered_by_id_CanIf(xdm_file,order_var,parent):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()
        root = etree.fromstring(xdm_content)
        elements = root.xpath(f".//d:lst[@name='{parent}']/d:ctr", namespaces=namespace)
        frames_data = [(ctr.attrib['name'], ctr.xpath(f"string(d:var[@name='{order_var}']/@value)", namespaces=namespace)) for ctr in elements]
        frames_data = [(name, Id) for name, Id in frames_data if Id.strip()]
        first_Id = int(frames_data[0][1])
        if first_Id != 0:
            return "The first frame's ("+order_var+") should be (0)', but found ("+str(first_Id)+")."

        Ids = [int(Id) for _, Id in frames_data]
        if len(Ids) != len(set(Ids)):
            duplicates = [frame_name for frame_name, Id in frames_data if Ids.count(int(Id)) > 1]
            errorstring=""
            for frame_name in duplicates:
                errorstring=errorstring+' '+"The frame ("+frame_name+") has a duplicate ("+order_var+").\n"
            return errorstring

        Last_Id = int(frames_data[-1][1])
        total_frames = len(frames_data)
        if Last_Id != total_frames - 1:
            return "The last frame's ("+order_var+") should be ("+str(total_frames-1)+"), but found ("+str(Last_Id)+")."

        if any(int(frames_data[i - 1][1]) > int(frames_data[i][1]) for i in range(1, len(frames_data))):
            frame_name = frames_data[next(i for i in range(1, len(frames_data)) if int(frames_data[i - 1][1]) > int(frames_data[i][1]))][0]
            return "The frame ("+frame_name+") has a jump in ("+order_var+")."

        return True

    except Exception as e:
        print(f"Error occurred while processing the CANIF file: {e}")
        return False

#Ordering by index table PDUR
def ordered_by_id_PDUR(xdm_file,nodes):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()

        root = etree.fromstring(xdm_content)
        if(nodes[2]=='PduRDestPdu'):
            elements = root.xpath(f".//d:ctr[@name='{nodes[0]}']/d:lst[@name='{nodes[1]}']/d:ctr/d:lst[@name='{nodes[2]}']/d:ctr", namespaces=namespace)
            frames_data = [(ctr.attrib['name'],ctr.xpath(f".//d:var[@name='{nodes[3]}']/@value", namespaces=namespace)[0]) for ctr in elements]
        else:
            elements = root.xpath(f".//d:ctr[@name='{nodes[0]}']/d:lst[@name='{nodes[1]}']/d:ctr", namespaces=namespace)
            frames_data = [(ctr.attrib['name'], ctr.xpath(f".//d:ctr[@name='{nodes[2]}']/d:var[@name='{nodes[3]}']/@value", namespaces=namespace)[0]) for ctr in elements]

        frames_data = [(name, Id) for name, Id in frames_data if Id.strip()]
        first_Id = int(frames_data[0][1])
        if first_Id != 0:
            return "The first frame's ("+nodes[3]+") should be (0), but found ("+str(first_Id)+")."

        Ids = [int(Id) for _, Id in frames_data]
        if len(Ids) != len(set(Ids)):
            duplicates = [frame_name for frame_name, Id in frames_data if Ids.count(int(Id)) > 1]
            errorstring=""
            for frame_name in duplicates:
                errorstring=errorstring+"The frame ("+frame_name+") has a duplicate "+nodes[3]+".\n"
            return errorstring

        Last_Id = int(frames_data[-1][1])
        total_frames = len(frames_data)
        if nodes[2]=='PduRDestPdu':
            if Last_Id != total_frames:
                return "The last frame's ("+order_var+") should be ("+str(total)+"), but found ("+str(Last_Id)+")."
        else:
            if Last_Id != total_frames-1:
                return "The last frame's ("+order_var+") should be ("+str(total-1)+"), but found ("+str(Last_Id)+")."
                
        if any(int(frames_data[i - 1][1]) > int(frames_data[i][1]) for i in range(1, len(frames_data))):
            frame_name = frames_data[next(i for i in range(1, len(frames_data)) if int(frames_data[i - 1][1]) > int(frames_data[i][1]))][0]
            return "The frame ("+frame_name+") has a jump in "+nodes[3]+"."

        return True

    except Exception as e:
        print(f"Error occurred while processing the PDUR file: {e}")
        return False

#Ordering by ComIPduHandleId or ComHandleId check
def ordered_by_id_COM(xdm_file,order_var,parent):
    try:
        data_type=""
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()
        root = etree.fromstring(xdm_content)
        if(order_var=="ComIPduHandleId"):
            data_type="frame"
        else:
            data_type="signal"
        elements = root.xpath(f".//d:lst[@name='{parent}']/d:ctr", namespaces=namespace)
        data = [(ctr.attrib['name'], ctr.xpath(f"string(d:var[@name='{order_var}']/@value)", namespaces=namespace)) for ctr in elements]
        data = [(name, Id) for name, Id in data if Id.strip()]
        first_Id = int(data[0][1])
        if data_type=="frame":
            if first_Id != 0:
                return "The first "+data_type+"'s ("+order_var+") should be (0), but found ("+str(first_Id)+")."
        else:
            if first_Id != 1:
                return "The first "+data_type+"'s ("+order_var+") should be (1), but found ("+str(first_Id)+")."
        Ids = [int(Id) for _, Id in data]
        if len(Ids) != len(set(Ids)):
            duplicates = [name for name, Id in data if Ids.count(int(Id)) > 1]
            errorstring=""
            for name in duplicates:
                errorstring=errorstring+' '+"The "+data_type+" ("+name+") has a duplicate ("+order_var+").\n"
            return errorstring

        Last_Id = int(data[-1][1])
        total = len(data)
        if data_type=="frame":
            if Last_Id != total - 1:
                return "The last "+data_type+"'s ("+order_var+") should be ("+str(total-1)+"), but found ("+str(Last_Id)+")."
        else:
            if Last_Id != total:
                return "The last "+data_type+"'s ("+order_var+") should be ("+str(total)+"), but found ("+str(Last_Id)+")."
        if any(int(data[i - 1][1]) > int(data[i][1]) for i in range(1, len(data))):
            name = data[next(i for i in range(1, len(data)) if int(data[i - 1][1]) > int(data[i][1]))][0]
            return "The "+data_type+" ("+name+") has a jump in ("+order_var+")."

        return True

    except Exception as e:
        print(f"Error occurred while processing the COM file: {e}")
        return False

#Computes the signal's position in the frame
def Signal_Position_inFrame(_position_signal_dbc_, _taille_signal_dbc_):
    
    DIV_ENT = int(_position_signal_dbc_) // 8    
    REST1 = (DIV_ENT + 1) * 8 - int(_position_signal_dbc_)
    REST2 = int(_taille_signal_dbc_) - REST1
    residu = REST2 % 8   
    if residu == 0 :
       k = (REST2 // 8) - 1
    else:
       k = REST2 // 8
    
    if REST2 > 0 :
       _pos_calc_cdc_ = (DIV_ENT - 1 - k) *8 + REST2 - 8 * k - 1
    else:
       _pos_calc_cdc_ = int(_position_signal_dbc_) + int(_taille_signal_dbc_) - 1  
    _Signal_Byte_Position_CDC_ = ( _pos_calc_cdc_ // 8 ) + 1
    _Signal_Bit_Position_CDC_ = _pos_calc_cdc_ % 8

    return _Signal_Byte_Position_CDC_, _Signal_Bit_Position_CDC_

#determines the type of the signal based on the length ,val max,val min,resolution and offset
def Signal_Type(_longueur_, _Signal_phy_min_, _Signal_phy_max_, _Signal_Phy_Resolution_, _Signal_Offset_):
    if int(float(_longueur_))  == 1:
       _type_ = "BOOLEAN"
    elif _Signal_phy_min_ != "Non applicable" or _Signal_phy_max_ != "Non applicable" :
 
        SUP = (float(_Signal_phy_max_) - float(_Signal_Offset_)) // float (_Signal_Phy_Resolution_)
        INF = (float(_Signal_phy_min_) - float(_Signal_Offset_)) // float (_Signal_Phy_Resolution_)
        if (SUP >=0 and  SUP <= 255)  and (INF >=0 and  INF <= 255):
           _type_ = "UINT8"
        elif SUP <= 127 and  INF >= -128:
           _type_ = "SINT8"	
        elif (SUP >=0 and  SUP <= 65535)  and (INF >=0 and  INF <= 65535):	
           _type_ = "UINT16"	   
        elif SUP <= 32767 and  INF >= (-32768):
           _type_ = "SINT16"
        elif (SUP >=0 and  SUP <= 4294967295)  and (INF >=0 and  INF <= 4294967295):			  
           _type_ = "UINT32"		  
        elif SUP <= 2147483647 and  INF >= (-2147483648):
           _type_ = "SINT32"	
    elif int(float(_longueur_))  <= 8:
       _type_ = "UINT8"
    elif int(float(_longueur_))  <= int(16):
       _type_ = "UINT16"	
    elif int(float(_longueur_))  <= int(32):
       _type_ = "UINT32"
    elif int(float(_longueur_))  <= int(64):
       _type_ = "UINT64"	   
    return _type_
