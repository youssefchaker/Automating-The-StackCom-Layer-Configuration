import pandas as pd
import tkinter as tk
import os
from tkinter import filedialog
from lxml import etree
from openpyxl import load_workbook

file_path = os.path.join(os.getcwd(), 'Output.xlsx')

# Define expected headers for cleaning the Excel data
expected_headers = {'FRAMES': ['Radical', 'Activation trame', 'Protocole_M', 'Identifiant_T', 'Taille_Max_T', 'Lmin_T', 'Mode_Transmission_T', 'Nature_Evenement_FR_T', 'Nature_Evenement_GB_T', 'Periode_T', 'UCE Emetteur', 'AEE10r3 Reseau_T']}

# Function to clean the Excel data and keep only the necessary columns
def cleanExcelData(excel_file):
    df = pd.read_excel(excel_file, sheet_name='FRAMES', header=0)
    headers = [col for col in df.columns if col in expected_headers['FRAMES']]
    return df[headers]

# Function to write data to Excel file
def write_to_Excel(result_data, file_path):
    df = pd.DataFrame(result_data)

    if not os.path.exists(file_path):
        # Create the Excel file with the specified columns
        df.to_excel(file_path, sheet_name='CANIF_verif_PDU_Messagerie', index=False, header=True)
    else:
        # Load the existing workbook
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book

        if 'CANIF_verif_PDU_Messagerie' in pd.ExcelFile(file_path).sheet_names:
            sheet = book['CANIF_verif_PDU_Messagerie']
            # Append the data to the existing sheet
            df.to_excel(writer, sheet_name='CANIF_verif_PDU_Messagerie', index=False, header=False, startrow=writer.sheets['CANIF_verif_PDU_Messagerie'].max_row)

        else:
            # Create a new sheet if it doesn't exist
            df.to_excel(writer, sheet_name='CANIF_verif_PDU_Messagerie', index=False, header=True)

        writer.save()


# Function to extract necessary attributes for the target frame from the .xdm file
def extract_CanifValues(xdm_file, frame_name,excel_file_path):
    with open(xdm_file, 'r') as file:
        xdm_content = file.read()
    
    frames_data = cleanExcelData(excel_file_path)

    root = etree.fromstring(xdm_content)
    namespace = {'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'}
    selected_frame = frames_data[frames_data['Radical'] == frame_name]
    if selected_frame.empty:
        result_data = {
            'Frame Name': [frame_name],
            'Passed?': ["Frame Not Found in Messagerie "],
            'Frame type':' ',
            'CanIfCanCtrlIdRef':' ',
            'AEE10r3 Reseau_T': ' ',
            'CanIfCanCtrlIdRef/AEE10r3 Reseau_T':' ',
            'CanIfCanHandleTypeRef':' ',
            'CanIfCanHandleTypeRef/Frame Name': ' ',
            'CanIfIdSymRef':' ',
            'CanIfIdSymRef/Frame Name':' ',
        }
        write_to_Excel(result_data,file_path)
        CanIfCanCtrlIdRef, CanIfCanHandleTypeRef, CanIfIdSymRef= -1, -1, -1
        return CanIfCanCtrlIdRef, CanIfCanHandleTypeRef, CanIfIdSymRef
    else:
        if( not selected_frame["UCE Emetteur"].str.endswith("E_VCU").any()):
            ctr_elements = root.xpath(".//d:lst[@name='CanIfHrhCfg']/d:ctr[contains(@name, $name)]", namespaces=namespace, name=frame_name)
            if ctr_elements:
                CanIfCanCtrlIdRef = ctr_elements[0].xpath("string(d:ref[@name='CanIfHrhCanCtrlIdRef']/@value)", namespaces=namespace)
                CanIfCanHandleTypeRef = ctr_elements[0].xpath("string(d:ref[@name='CanIfHrhCanHandleTypeRef']/@value)", namespaces=namespace)
                CanIfIdSymRef = ctr_elements[0].xpath("string(d:ref[@name='CanIfHrhIdSymRef']/@value)", namespaces=namespace)
            else:
                CanIfCanCtrlIdRef, CanIfCanHandleTypeRef, CanIfIdSymRef= None, None, None
        else:
            ctr_elements = root.xpath(".//d:lst[@name='CanIfHthCfg']/d:ctr[contains(@name, $name)]", namespaces=namespace, name=frame_name)
            if ctr_elements:
                CanIfCanCtrlIdRef = ctr_elements[0].xpath("string(d:ref[@name='CanIfHthCanCtrlIdRef']/@value)", namespaces=namespace)
                CanIfCanHandleTypeRef = ctr_elements[0].xpath("string(d:ref[@name='CanIfHthCanHandleTypeRef']/@value)", namespaces=namespace)
                CanIfIdSymRef = ctr_elements[0].xpath("string(d:ref[@name='CanIfHthIdSymRef']/@value)", namespaces=namespace)
            else:
                CanIfCanCtrlIdRef, CanIfCanHandleTypeRef, CanIfIdSymRef= None, None, None
        return CanIfCanCtrlIdRef, CanIfCanHandleTypeRef, CanIfIdSymRef
           

def verify_frame(excel_file_path, xdm_file_path, frame_name):
    try:
        CanIfCanCtrlIdRef, CanIfCanHandleTypeRef, CanIfIdSymRef = extract_CanifValues(xdm_file_path, frame_name,excel_file_path)
        if CanIfCanCtrlIdRef== -1 and CanIfCanHandleTypeRef== -1 and CanIfIdSymRef== -1 :
            return
        elif CanIfCanCtrlIdRef is None and CanIfCanHandleTypeRef is None and CanIfIdSymRef is None :
            result_data = {
                    'Frame Name': [frame_name],
                    'Passed?':["Frame Not Found in CANIF"],
                    'Frame type':' ',
                    'CanIfCanCtrlIdRef':' ',
                    'AEE10r3 Reseau_T': ' ',
                    'CanIfCanCtrlIdRef/AEE10r3 Reseau_T':' ',
                    'CanIfCanHandleTypeRef':' ',
                    'CanIfCanHandleTypeRef/Frame Name': ' ',
                    'CanIfIdSymRef':' ',
                    'CanIfIdSymRef/Frame Name':' ',
                }
            write_to_Excel(result_data,file_path)
            return False
        else:
            frames_data = cleanExcelData(excel_file_path)
            selected_frame = frames_data[frames_data['Radical'] == frame_name]
            CanIfCanCtrlIdReftst=CanIfCanHandleTypeReftst=CanIfIdSymReftst=True

            if "ASPath:/CanIf/CanIf/CanIfCtrlDrvCfg/CanIf_Controller_Can_InterSystem" == CanIfCanCtrlIdRef and selected_frame["AEE10r3 Reseau_T"].values[0].startswith("HS1"):
                pass
            elif "ASPath:/CanIf/CanIf/CanIfCtrlDrvCfg/CanIf_Controller_Can_LAS" == CanIfCanCtrlIdRef and selected_frame["AEE10r3 Reseau_T"].values[0].startswith("HS2"):
                pass
            elif "ASPath:/CanIf/CanIf/CanIfCtrlDrvCfg/CanIf_Controller_Can_eCAN" == CanIfCanCtrlIdRef and selected_frame["AEE10r3 Reseau_T"].values[0].startswith("E_CAN"):
                pass
            else:
                    CanIfCanCtrlIdReftst=False
            
            if(frame_name not in CanIfCanHandleTypeRef):
                CanIfCanHandleTypeReftst=False
            if(frame_name not in CanIfIdSymRef):
                CanIfIdSymReftst=False
            result_data = {
                'Frame Name': [frame_name],
                'Passed?':[" " if CanIfCanCtrlIdReftst ==False or CanIfIdSymReftst==False or CanIfCanHandleTypeReftst==False else "X"],
                'Frame type':["TRANSMIT" if selected_frame["UCE Emetteur"].str.endswith("E_VCU").any() else "RECEIVE" ],
                'CanIfCanCtrlIdRef':[CanIfCanCtrlIdRef],
                'AEE10r3 Reseau_T': [selected_frame["AEE10r3 Reseau_T"].values[0]],
                'CanIfCanCtrlIdRef/AEE10r3 Reseau_T Errors': ['Error (CanIfCanCtrlIdRef Mismatch)' if CanIfCanCtrlIdReftst==False else "None"],
                'CanIfCanHandleTypeRef':[CanIfCanHandleTypeRef],
                'CanIfCanHandleTypeRef/Frame Name Errors': ['Error (Frame Name not present in CanIfCanHandleTypeRef)' if CanIfCanHandleTypeReftst==False else "None"],
                'CanIfIdSymRef':[CanIfIdSymRef],
                'CanIfIdSymRef/Frame Name Errors': ['Error (Frame Name not present in CanIfIdSymRef)' if CanIfIdSymReftst==False else "None"],
            }
            write_to_Excel(result_data,file_path)

    except Exception as e:
                print(f"Error occurred : {e}")
                return False     


# Clear the Excel file
def clear_excel():

    # Create an Excel writer object
    excel_writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
    df = pd.DataFrame(columns=['Frame Name',
                    'Passed?',
                    'Frame Type',
                    'CanIfCanCtrlIdRef',
                    'AEE10r3 Reseau_T',
                    'CanIfCanCtrlIdRef/AEE10r3 Reseau_T',
                    'CanIfCanHandleTypeRef',
                    'CanIfCanHandleTypeRef/Frame Name',
                    'CanIfIdSymRef',
                    'CanIfIdSymRef/Frame Name'])
    df.to_excel(excel_writer, sheet_name='CANIF_verif_PDU_Messagerie', index=False)
    excel_writer.save()
    completion_label.config(text="Output File Cleared", fg="blue")


#select the excel file from the interface
def browse_excel():
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not excel_file_path:
        return
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(tk.END, excel_file_path)


#select the xdm file from the interface
def browse_xdm():
    xdm_file_path = filedialog.askopenfilename(filetypes=[("XDM files", "*.xdm")])
    if not xdm_file_path:
        return
    xdm_file_entry.delete(0, tk.END)
    xdm_file_entry.insert(tk.END, xdm_file_path)


#execute functionality on button click
def verify_button_click():
    excel_file_path = excel_file_entry.get()
    xdm_file_path = xdm_file_entry.get()
    frame_name = frame_entry.get()

    verify_frame(excel_file_path, xdm_file_path, frame_name)
    completion_label.config(text="Output Created", fg="green")


# Create the GUI
root = tk.Tk()
root.title("Frame Info CANIF/Messagerie Verification")

frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

excel_file_label = tk.Label(frame, text="Select Excel File:")
excel_file_label.grid(row=0, column=0, padx=5, pady=5)

excel_file_entry = tk.Entry(frame)
excel_file_entry.grid(row=0, column=1, padx=5, pady=5)

excel_file_button = tk.Button(frame, text="Browse", command=browse_excel)
excel_file_button.grid(row=0, column=2, padx=5, pady=5)

xdm_file_label = tk.Label(frame, text="Select Canif File:")
xdm_file_label.grid(row=1, column=0, padx=5, pady=5)

xdm_file_entry = tk.Entry(frame)
xdm_file_entry.grid(row=1, column=1, padx=5, pady=5)

xdm_file_button = tk.Button(frame, text="Browse", command=browse_xdm)
xdm_file_button.grid(row=1, column=2, padx=5, pady=5)

frame_label = tk.Label(frame, text="Enter Frame Name:")
frame_label.grid(row=2, column=0, padx=5, pady=5)

frame_entry = tk.Entry(frame)
frame_entry.grid(row=2, column=1, padx=5, pady=5)

verify_button = tk.Button(frame, text="Verify", command=verify_button_click)
verify_button.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

clear_excel_button = tk.Button(frame, text="Clear Excel", command=clear_excel)
clear_excel_button.grid(row=6, column=0, columnspan=3, padx=5, pady=5)

completion_label = tk.Label(frame, text="", fg="green")
completion_label.grid(row=7, column=0, columnspan=3, padx=5, pady=5)

root.mainloop()