import tkinter as tk
from tkinter import filedialog
from lxml import etree
import pandas as pd
import tkinter as tk
import os
from openpyxl import load_workbook

file_path = os.path.join(os.getcwd(), 'Output.xlsx')

def write_to_Excel(result_data, file_path):
    df = pd.DataFrame(result_data)

    if not os.path.exists(file_path):
        # Create the Excel file with the specified columns
        df.to_excel(file_path, sheet_name='PDUR_index_table_order_verif', index=False, header=True)
    else:
        # Load the existing workbook
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book

        if 'PDUR_index_table_order_verif' in pd.ExcelFile(file_path).sheet_names:
            # Check if the 'Passed?' column already exists in the sheet
            sheet = book['PDUR_index_table_order_verif']
            # Append the data to the existing sheet
            df.to_excel(writer, sheet_name='PDUR_index_table_order_verif', index=False, header=False, startrow=writer.sheets['PDUR_index_table_order_verif'].max_row)

        else:
            # Create a new sheet if it doesn't exist
            df.to_excel(writer, sheet_name='PDUR_index_table_order_verif', index=False, header=True)

        writer.save()

# Clear the Excel file
def clear_excel():
    sheet_name='PDUR_index_table_order_verif'
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            sheet.delete_rows(2, sheet.max_row)
        book.save(file_path)
    completion_label.config(text="Output File Cleared", fg="blue")

#Ordering by PduRSourcePduHandleId
def ordered_by_id_Tx_src(xdm_file):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()

        root = etree.fromstring(xdm_content)
        namespace = {'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd','a':'http://www.tresos.de/_projects/DataModel2/08/attribute.xsd'}
        ctr_elements = root.xpath(".//d:ctr[@name='Com_PduRRoutingTable']/d:lst[@name='PduRRoutingPath']/d:ctr", namespaces=namespace)
        frames_data = [(ctr.attrib['name'], ctr.xpath("string(.//d:ctr[@name='PduRSrcPdu']/d:var[@name='PduRSourcePduHandleId']/@value)", namespaces=namespace)) for ctr in ctr_elements]
        frames_data = [(name, PduRSourcePduHandleId) for name, PduRSourcePduHandleId in frames_data if PduRSourcePduHandleId.strip()]
        first_PduRSourcePduHandleId = int(frames_data[0][1])
        if first_PduRSourcePduHandleId != 0:
            return "The first frame's PduRSourcePduHandleId should be (0), but found ("+str(first_PduRSourcePduHandleId)+")"

        PduRSourcePduHandleIds = [int(PduRSourcePduHandleId) for _, PduRSourcePduHandleId in frames_data]
        if len(PduRSourcePduHandleIds) != len(set(PduRSourcePduHandleIds)):
            duplicates = [frame_name for frame_name, PduRSourcePduHandleId in frames_data if PduRSourcePduHandleIds.count(int(PduRSourcePduHandleId)) > 1]
            errorstring=""
            for frame_name in duplicates:
                errorstring=errorstring+"The frame ("+frame_name+") has a duplicate PduRSourcePduHandleId\n"
            return errorstring

        Last_PduRSourcePduHandleId = int(frames_data[-1][1])
        total_frames = len(frames_data)
        if Last_PduRSourcePduHandleId != total_frames - 1:
            return "The last frame's PduRSourcePduHandleId should be ("+str(total_frames-1)+"), but found ("+str(Last_PduRSourcePduHandleId)+")"

        if any(int(frames_data[i - 1][1]) > int(frames_data[i][1]) for i in range(1, len(frames_data))):
            frame_name = frames_data[next(i for i in range(1, len(frames_data)) if int(frames_data[i - 1][1]) > int(frames_data[i][1]))][0]
            return "The frame ("+frame_name+") has a jump in PduRSourcePduHandleId"

        return True

    except Exception as e:
        print(f"Error: {e}")
        return False

#Ordering by PduRDestPduHandleId
def ordered_by_id_Tx_dest(xdm_file):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()

        root = etree.fromstring(xdm_content)
        namespace = {'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd','a':'http://www.tresos.de/_projects/DataModel2/08/attribute.xsd'}
        ctr_elements = root.xpath(".//d:ctr[@name='Com_PduRRoutingTable']/d:lst[@name='PduRRoutingPath']/d:ctr/d:lst[@name='PduRDestPdu']/d:ctr", namespaces=namespace)
        frames_data = [(ctr.attrib['name'], ctr.xpath("string(.//d:var[@name='PduRDestPduHandleId']/@value)", namespaces=namespace)) for ctr in ctr_elements]
        frames_data = [(name, PduRDestPduHandleId) for name, PduRDestPduHandleId in frames_data if PduRDestPduHandleId.strip()]
        first_PduRDestPduHandleId = int(frames_data[0][1])
        if first_PduRDestPduHandleId != 0:
            return "The first frame's PduRDestPduHandleId should be (0), but found ("+str(first_PduRDestPduHandleId)+")"

        PduRDestPduHandleIds = [int(PduRDestPduHandleId) for _, PduRDestPduHandleId in frames_data]
        if len(PduRDestPduHandleIds) != len(set(PduRDestPduHandleIds)):
            duplicates = [frame_name for frame_name, PduRDestPduHandleId in frames_data if PduRDestPduHandleIds.count(int(PduRDestPduHandleId)) > 1]
            errorstring=""
            for frame_name in duplicates:
                errorstring=errorstring+"The frame ("+frame_name+") has a duplicate PduRDestPduHandleId\n"
            return errorstring

        Last_PduRDestPduHandleId = int(frames_data[-1][1])
        total_frames = len(frames_data)
        if Last_PduRDestPduHandleId != total_frames - 1:
            return "The last frame's PduRDestPduHandleId should be ("+str(total_frames-1)+"), but found ("+str(Last_PduRDestPduHandleId)+")"

        if any(int(frames_data[i - 1][1]) > int(frames_data[i][1]) for i in range(1, len(frames_data))):
            frame_name = frames_data[next(i for i in range(1, len(frames_data)) if int(frames_data[i - 1][1]) > int(frames_data[i][1]))][0]
            return "The frame ("+frame_name+") has a jump in PduRDestPduHandleId"

        return True

    except Exception as e:
        print(f"Error: {e}")
        return False


def ordered_by_id_Rx_src(xdm_file):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()

        root = etree.fromstring(xdm_content)
        namespace = {'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd','a':'http://www.tresos.de/_projects/DataModel2/08/attribute.xsd'}
        ctr_elements = root.xpath(".//d:ctr[@name='CanIf_PduRRoutingTable']/d:lst[@name='PduRRoutingPath']/d:ctr", namespaces=namespace)
        frames_data = [(ctr.attrib['name'], ctr.xpath("string(.//d:ctr[@name='PduRSrcPdu']/d:var[@name='PduRSourcePduHandleId']/@value)", namespaces=namespace)) for ctr in ctr_elements]
        frames_data = [(name, PduRSourcePduHandleId) for name, PduRSourcePduHandleId in frames_data if PduRSourcePduHandleId.strip()]
        first_PduRSourcePduHandleId = int(frames_data[0][1])
        if first_PduRSourcePduHandleId != 0:
            return "The first frame's PduRSourcePduHandleId should be (0), but found ("+str(first_PduRSourcePduHandleId)+")"

        PduRSourcePduHandleIds = [int(PduRSourcePduHandleId) for _, PduRSourcePduHandleId in frames_data]
        if len(PduRSourcePduHandleIds) != len(set(PduRSourcePduHandleIds)):
            duplicates = [frame_name for frame_name, PduRSourcePduHandleId in frames_data if PduRSourcePduHandleIds.count(int(PduRSourcePduHandleId)) > 1]
            errorstring=""
            for frame_name in duplicates:
                errorstring=errorstring+"The frame ("+frame_name+") has a duplicate PduRSourcePduHandleId\n"
            return errorstring

        Last_PduRSourcePduHandleId = int(frames_data[-1][1])
        total_frames = len(frames_data)
        if Last_PduRSourcePduHandleId != total_frames - 1:
            return "The last frame's PduRSourcePduHandleId should be ("+str(total_frames-1)+"), but found ("+str(Last_PduRSourcePduHandleId)+")"

        if any(int(frames_data[i - 1][1]) > int(frames_data[i][1]) for i in range(1, len(frames_data))):
            frame_name = frames_data[next(i for i in range(1, len(frames_data)) if int(frames_data[i - 1][1]) > int(frames_data[i][1]))][0]
            return "The frame ("+frame_name+") has a jump in PduRSourcePduHandleId"

        return True

    except Exception as e:
        print(f"Error: {e}")
        return False

#Ordering by PduRDestPduHandleId
def ordered_by_id_Rx_dest(xdm_file):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()

        root = etree.fromstring(xdm_content)
        namespace = {'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd','a':'http://www.tresos.de/_projects/DataModel2/08/attribute.xsd'}
        ctr_elements = root.xpath(".//d:ctr[@name='CanIf_PduRRoutingTable']/d:lst[@name='PduRRoutingPath']/d:ctr/d:lst[@name='PduRDestPdu']/d:ctr", namespaces=namespace)
        frames_data = [(ctr.attrib['name'], ctr.xpath("string(.//d:var[@name='PduRDestPduHandleId']/@value)", namespaces=namespace)) for ctr in ctr_elements]
        frames_data = [(name, PduRDestPduHandleId) for name, PduRDestPduHandleId in frames_data if PduRDestPduHandleId.strip()]
        first_PduRDestPduHandleId = int(frames_data[0][1])
        if first_PduRDestPduHandleId != 0:
            return "The first frame's PduRDestPduHandleId should be (0), but found ("+str(first_PduRDestPduHandleId)+")"

        PduRDestPduHandleIds = [int(PduRDestPduHandleId) for _, PduRDestPduHandleId in frames_data]
        if len(PduRDestPduHandleIds) != len(set(PduRDestPduHandleIds)):
            duplicates = [frame_name for frame_name, PduRDestPduHandleId in frames_data if PduRDestPduHandleIds.count(int(PduRDestPduHandleId)) > 1]
            errorstring=""
            for frame_name in duplicates:
                errorstring=errorstring+"The frame ("+frame_name+") has a duplicate PduRDestPduHandleId\n"
            return errorstring

        Last_PduRDestPduHandleId = int(frames_data[-1][1])
        total_frames = len(frames_data)
        if Last_PduRDestPduHandleId != total_frames - 1:
            return "The last frame's PduRDestPduHandleId should be ("+str(total_frames-1)+"), but found ("+str(Last_PduRDestPduHandleId)+")"

        if any(int(frames_data[i - 1][1]) > int(frames_data[i][1]) for i in range(1, len(frames_data))):
            frame_name = frames_data[next(i for i in range(1, len(frames_data)) if int(frames_data[i - 1][1]) > int(frames_data[i][1]))][0]
            return "The frame ("+frame_name+") has a jump in PduRDestPduHandleId"

        return True

    except Exception as e:
        print(f"Error: {e}")
        return False


def check_order():
    xdm_file_path = xdm_file_entry.get()
    if not xdm_file_path:
        return
    result_data = {
        'Passed?':["X" if ordered_by_id_Tx_src(xdm_file_path)==True and ordered_by_id_Rx_src(xdm_file_path)==True and ordered_by_id_Tx_dest(xdm_file_path)==True and ordered_by_id_Rx_dest(xdm_file_path)==True else " "],
        'Order by Tx_PduRSourcePduHandleId':[" " if ordered_by_id_Tx_src(xdm_file_path)==True else ordered_by_id_Tx_src(xdm_file_path)],
        'Order by Tx_PduRDestPduHandleId':[" " if ordered_by_id_Tx_dest(xdm_file_path)==True else ordered_by_id_Tx_dest(xdm_file_path)],
        'Order by Rx_PduRSourcePduHandleId':[" " if ordered_by_id_Rx_src(xdm_file_path)==True else ordered_by_id_Rx_src(xdm_file_path)],
        'Order by Rx_PduRDestPduHandleId':[" " if ordered_by_id_Rx_dest(xdm_file_path)==True else ordered_by_id_Rx_dest(xdm_file_path)]
        
     }
    write_to_Excel(result_data,file_path)
    completion_label.config(text="Output Created", fg="green")

    
#open the xdm file
def browse_pdur():
    xdm_file_path = filedialog.askopenfilename(filetypes=[("XDM files", "*XDM")])
    if not xdm_file_path:
        return
    xdm_file_entry.delete(0, tk.END)
    xdm_file_entry.insert(tk.END, xdm_file_path)

root = tk.Tk()
root.title("PDUR.xdm File Order Checker")

frame = tk.Frame(root, padx=10, pady=10)
frame.pack()

xml_file_label = tk.Label(frame, text="Select PDUR File:")
xml_file_label.grid(row=0, column=0)

xdm_file_entry = tk.Entry(frame)
xdm_file_entry.grid(row=0, column=1)

xdm_file_button = tk.Button(frame, text="Browse", command=browse_pdur)
xdm_file_button.grid(row=0, column=2)

check_receive_transmit_button = tk.Button(frame, text="Check Order", command=check_order)
check_receive_transmit_button.grid(row=1, column=0, columnspan=3, pady=5)

clear_excel_button = tk.Button(frame, text="Clear Excel", command=clear_excel)
clear_excel_button.grid(row=2, column=0, columnspan=3, pady=5)

completion_label = tk.Label(frame, text="", fg="green")
completion_label.grid(row=7, column=0, columnspan=3, padx=5, pady=5)

root.mainloop()
