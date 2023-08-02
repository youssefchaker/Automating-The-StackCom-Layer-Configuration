import tkinter as tk
from tkinter import filedialog
from lxml import etree
import logging
import pandas as pd
import tkinter as tk
import os
from openpyxl import load_workbook

file_path = os.path.join(os.getcwd(), 'Output.xlsx')

def write_to_Excel(result_data, file_path):
    df = pd.DataFrame(result_data)

    if not os.path.exists(file_path):
        # Create the Excel file with the specified columns
        df.to_excel(file_path, sheet_name='CAN_verif_Geeds', index=False, header=True)
    else:
        # Load the existing workbook
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book

        if 'CAN_verif_Geeds' in pd.ExcelFile(file_path).sheet_names:
            # Check if the 'Passed?' column already exists in the sheet
            sheet = book['CAN_verif_Geeds']
            # Append the data to the existing sheet
            df.to_excel(writer, sheet_name='CAN_verif_Geeds', index=False, header=False, startrow=writer.sheets['CAN_verif_Geeds'].max_row)

        else:
            # Create a new sheet if it doesn't exist
            df.to_excel(writer, sheet_name='CAN_verif_Geeds', index=False, header=True)

        writer.save()

# Clear the Excel file
def clear_excel():
    sheet_name='CAN_verif_Geeds'
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            sheet.delete_rows(2, sheet.max_row)
        book.save(file_path)
    completion_label.config(text="Output File Cleared", fg="blue")

#Ordering by TRANSMIT and RECIEVE
def ordered_by_RX_TX(xdm_file):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()

        root = etree.fromstring(xdm_content)
        ctr_elements = root.xpath(".//d:lst[@name='CanHardwareObject']/d:ctr", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'})

        receive_indices = [i for i, ctr in enumerate(ctr_elements) if ctr.xpath("string(d:var[@name='CanObjectType']/@value)", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'}) == "RECEIVE"]
        transmit_indices = [i for i, ctr in enumerate(ctr_elements) if ctr.xpath("string(d:var[@name='CanObjectType']/@value)", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'}) == "TRANSMIT"]

        if receive_indices and transmit_indices:
            if receive_indices[-1] > transmit_indices[0]:
                frame_name = ctr_elements[transmit_indices[0]].attrib['name']
                return "TRANSMIT frame("+frame_name+") before RECEIVE frames"

            if any(ctr.xpath("string(d:var[@name='CanObjectType']/@value)", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'}) == "RECEIVE" for ctr in ctr_elements[transmit_indices[-1] + 1:]):
                frame_name = ctr_elements[transmit_indices[0]].attrib['name']
                return "TRANSMIT frame("+frame_name+"}) before RECEIVE frames"

        return True

    except Exception as e:
        print(f"Error occurred while processing the XDM file: {e}")
        return False

#Ordering by Cancontrollerref
def ordered_by_CAN_Ref(xdm_file):
    expected_order = ['CAN_2', 'CAN_1', 'CAN_DEVAID', 'CAN_3']
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()

        root = etree.fromstring(xdm_content)
        ctr_elements = root.xpath(".//d:lst[@name='CanHardwareObject']/d:ctr", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'})

        def check_order(frames):
            prev_index = None
            for ctr in frames:
                ref_value = ctr.xpath("string(d:ref[@name='CanControllerRef']/@value)", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'})
                index = next((i for i, val in enumerate(expected_order) if val in ref_value), None)
                if index is not None:
                    if prev_index is not None and index < prev_index:
                        frame_name = ctr.attrib['name']
                        return "The frame ("+frame_name+") has incorrect 'CanControllerRef' attribute order should be ("+expected_order[index]+")}"
                    prev_index = index
                else:
                    frame_name = ctr.attrib['name']
                    return "The frame ("+frame_name+") has incorrect 'CanControllerRef' attribute order should be ("+expected_order[index]+")"
            return "True"

        receive_frames = [ctr for ctr in ctr_elements if ctr.xpath("string(d:var[@name='CanObjectType']/@value)", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'}) == "RECEIVE"]
        transmit_frames = [ctr for ctr in ctr_elements if ctr.xpath("string(d:var[@name='CanObjectType']/@value)", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'}) == "TRANSMIT"]

        receive_order_check = check_order(receive_frames)
        transmit_order_check = check_order(transmit_frames)

        if(receive_order_check!="True" and transmit_order_check!="True" ):
            return receive_order_check+transmit_order_check
        elif (receive_order_check!="True"):
            return receive_order_check
        elif(transmit_order_check!="True"):
            return transmit_order_check
        else:
            return True

    except Exception as e:
        print(f"Error occurred while processing the XDM file: {e}")
        return False


#Ordering by CanObjectId
def ordered_by_id(xdm_file):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()

        root = etree.fromstring(xdm_content)
        ctr_elements = root.xpath(".//d:lst[@name='CanHardwareObject']/d:ctr", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'})

        frames_data = [(ctr.attrib['name'], ctr.xpath("string(d:var[@name='CanObjectId']/@value)", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'})) for ctr in ctr_elements]
        frames_data = [(name, obj_id) for name, obj_id in frames_data if obj_id.strip()]

        first_can_object_id = int(frames_data[0][1])
        if first_can_object_id != 0:
            return "The first frame's CanObjectId should be '0', but found '{first_can_object_id}'"

        can_object_ids = [int(obj_id) for _, obj_id in frames_data]
        if len(can_object_ids) != len(set(can_object_ids)):
            duplicates = [frame_name for frame_name, obj_id in frames_data if can_object_ids.count(int(obj_id)) > 1]
            errorstring=""
            for frame_name in duplicates:
                errorstring=errorstring+"The frame ("+frame_name+") has a duplicate CanObjectId\n"
            return errorstring

        last_can_object_id = int(frames_data[-1][1])
        total_frames = len(frames_data)
        if last_can_object_id != total_frames - 1:
            return "The last frame's CanObjectId should be '{total_frames - 1}', but found '{last_can_object_id}'"

        if any(int(frames_data[i - 1][1]) > int(frames_data[i][1]) for i in range(1, len(frames_data))):
            frame_name = frames_data[next(i for i in range(1, len(frames_data)) if int(frames_data[i - 1][1]) > int(frames_data[i][1]))][0]
            return "The frame ("+frame_name+") has a jump in CanObjectId"

        return True

    except Exception as e:
        print(f"Error occurred while processing the XDM file: {e}")
        return False

def check_all():
    xdm_file_path = xdm_file_entry.get()
    if not xdm_file_path:
        return
    result_data = {
        'Passed?':["X" if ordered_by_CAN_Ref(xdm_file_path)==ordered_by_id(xdm_file_path)==ordered_by_RX_TX(xdm_file_path) else " "],
        'Order by RX_TX':[" " if ordered_by_RX_TX(xdm_file_path)==True else ordered_by_RX_TX(xdm_file_path)],
        'Order by CanControllerRef':[" " if ordered_by_CAN_Ref(xdm_file_path)==True else ordered_by_CAN_Ref(xdm_file_path)],
        'Order by CanObjectId':[" " if ordered_by_id(xdm_file_path)==True else ordered_by_id(xdm_file_path)]
     }
    write_to_Excel(result_data,file_path)
    completion_label.config(text="Output Created", fg="green")

    
#open the xdm file
def browse_xdm():
    xdm_file_path = filedialog.askopenfilename(filetypes=[("XDM files", "*XDM")])
    if not xdm_file_path:
        return
    xdm_file_entry.delete(0, tk.END)
    xdm_file_entry.insert(tk.END, xdm_file_path)

root = tk.Tk()
root.title("Can.xdm File Order Checker")

frame = tk.Frame(root, padx=10, pady=10)
frame.pack()

xml_file_label = tk.Label(frame, text="Select CAN File:")
xml_file_label.grid(row=0, column=0)

xdm_file_entry = tk.Entry(frame)
xdm_file_entry.grid(row=0, column=1)

xml_file_button = tk.Button(frame, text="Browse", command=browse_xdm)
xml_file_button.grid(row=0, column=2)

check_receive_transmit_button = tk.Button(frame, text="Check Order", command=check_all)
check_receive_transmit_button.grid(row=1, column=0, columnspan=3, pady=5)

clear_excel_button = tk.Button(frame, text="Clear Excel", command=clear_excel)
clear_excel_button.grid(row=2, column=0, columnspan=3, pady=5)

completion_label = tk.Label(frame, text="", fg="green")
completion_label.grid(row=7, column=0, columnspan=3, padx=5, pady=5)

root.mainloop()
