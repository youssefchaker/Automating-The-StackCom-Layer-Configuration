import pandas as pd
import os
from openpyxl import load_workbook
from lxml import etree
import tkinter as tk
from tkinter import filedialog
import math

#the excel output file file path
file_path = os.path.join(os.getcwd(), 'Output.xlsx')

#the namespace used for the XDM documents
namespace = {'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd','a':'http://www.tresos.de/_projects/DataModel2/08/attribute.xsd'}

# Define expected headers for cleaning the Excel data
expected_trame_headers = {'FRAMES': ['Checked','Radical', 'Identifiant_T', 'Taille_Max_T', 'Lmin_T', 'Mode_Transmission_T', 'Periode_T', 'UCE Emetteur', 'AEE10r3 Reseau_T']}

# Function to clean the Excel data and keep only the necessary columns
def cleanExcelFrameData(excel_file):
    df = pd.read_excel(excel_file, sheet_name='FRAMES', header=None, skiprows=2)  # Skip first two rows
    df.columns = df.iloc[0]  # Use the third row as column names
    df = df.iloc[1:]  # Drop the second row (previous header)
    
    headers = [col for col in df.columns if col in expected_trame_headers['FRAMES']]
    df = df[headers]
    
    # Filter rows based on 'checked' column
    df = df[df['Checked'] == 'X']
    
    return df

# Define expected headers for cleaning the Excel data
expected_signal_headers = {'SIGNALS': ['Radical_T', 'Position_octet_S', 'Position_bit_S', 'Taille_Max_S', 'Mnemonique_S', 'Type_S', 'Valeur_Min_S', 'Valeur_Max_S', 'Resolution_S', 'Offset_S', 'Valeur_Invalide_S', 'Valeur_Indisponible_S', 'Valeur_Interdite_S','PROD_INIT','CONS_INIT' ,'Emetteur','Nécessité de sécurisation par checksum et compteur de process']}

# Function to clean the Excel data and keep only the necessary columns
def cleanExcelSignalData(excel_file):
    df = pd.read_excel(excel_file, sheet_name='SIGNALS', header=None, skiprows=2)  # Skip first two rows
    df.columns = df.iloc[0]  # Use the third row as column names
    df = df.iloc[1:]  # Drop the second row (previous header)
    
    headers = [col for col in df.columns if col in expected_signal_headers['SIGNALS']]
    df = df[headers]
    
    return df


#function responsible for writiting the output to the excel file
def write_to_Excel(result_data, file_path,sheet_name):
    df = pd.DataFrame(result_data)

    if not os.path.exists(file_path):
        # Create the Excel file with the specified columns
        df.to_excel(file_path, sheet_name=sheet_name, index=False, header=True)
    else:
        # Load the existing workbook
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book

        if sheet_name in pd.ExcelFile(file_path).sheet_names:
            # Check if the 'Passed?' column already exists in the sheet
            sheet = book[sheet_name]
            # Append the data to the existing sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=writer.sheets[sheet_name].max_row)

        else:
            # Create a new sheet if it doesn't exist
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        writer.save()

# function responsable for emptying the designated excel sheet
def clear_excel(sheet_name):
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            sheet.delete_rows(2, sheet.max_row - 1) 
        book.save(file_path)
    

#Ordering by index table CanIF
def ordered_by_id_CanIf(xdm_file,order_var,parent):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()
        root = etree.fromstring(xdm_content)
        elements = root.xpath(f".//d:lst[@name='{parent}']/d:ctr", namespaces=namespace)
        frames_data = [(ctr.attrib['name'], ctr.xpath(f"string(d:var[@name='{order_var}']/@value)", namespaces=namespace)) for ctr in elements]
        frames_data = [(name, Id) for name, Id in frames_data if Id.strip()]
        first_Id = int(frames_data[0][1])
        if first_Id != 0:
            return "The first frame's ("+order_var+") should be (0)', but found ("+str(first_Id)+")."

        Ids = [int(Id) for _, Id in frames_data]
        if len(Ids) != len(set(Ids)):
            duplicates = [frame_name for frame_name, Id in frames_data if Ids.count(int(Id)) > 1]
            errorstring=""
            for frame_name in duplicates:
                errorstring=errorstring+" "+"The frame ("+frame_name+") has a duplicate ("+order_var+").\n"
            return errorstring

        Last_Id = int(frames_data[-1][1])
        total_frames = len(frames_data)
        if Last_Id != total_frames - 1:
            return "The last frame's ("+order_var+") should be ("+str(total_frames-1)+"), but found ("+str(Last_Id)+")."

        if any(int(frames_data[i - 1][1]) > int(frames_data[i][1]) for i in range(1, len(frames_data))):
            frame_name = frames_data[next(i for i in range(1, len(frames_data)) if int(frames_data[i - 1][1]) > int(frames_data[i][1]))][0]
            return "The frame ("+frame_name+") has a jump in ("+order_var+")."

        return True

    except Exception as e:
        print(f"Error: {e}")
        return False

#Ordering by index table PDUR
def ordered_by_id_PDUR(xdm_file,nodes):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()

        root = etree.fromstring(xdm_content)
        if(nodes[2]=='PduRDestPdu'):
            ctr_elements = root.xpath(f".//d:ctr[@name='{nodes[0]}']/d:lst[@name='{nodes[1]}']/d:ctr/d:lst[@name='{nodes[2]}']/d:ctr", namespaces=namespace)
            frames_data = [(ctr.attrib['name'], ctr.xpath(f"string(.//d:var[@name='{nodes[3]}']/@value)", namespaces=namespace)) for ctr in ctr_elements]
        else:
            elements = root.xpath(f".//d:ctr[@name='{nodes[0]}']/d:lst[@name='{nodes[1]}']/d:ctr", namespaces=namespace)
            frames_data = [(ctr.attrib['name'], ctr.xpath(f"string(.//d:ctr[@name='{nodes[2]}']/d:var[@name='{nodes[3]}']/@value)", namespaces=namespace)) for ctr in elements]

        frames_data = [(name, Id) for name, Id in frames_data if Id.strip()]
        first_Id = int(frames_data[0][1])
        if first_Id != 0:
            return "The first frame's ("+nodes[3]+") should be (0), but found ("+str(first_Id)+")."

        Ids = [int(Id) for _, Id in frames_data]
        if len(Ids) != len(set(Ids)):
            duplicates = [frame_name for frame_name, Id in frames_data if Ids.count(int(Id)) > 1]
            errorstring=""
            for frame_name in duplicates:
                errorstring=errorstring+"The frame ("+frame_name+") has a duplicate ("+nodes[3]+").\n"
            return errorstring

        Last_Id = int(frames_data[-1][1])
        total_frames = len(frames_data)
        if Last_Id != total_frames - 1:
            return "The last frame's ("+nodes[3]+") should be ("+str(total_frames-1)+"), but found ("+str(Last_Id)+")."

        if any(int(frames_data[i - 1][1]) > int(frames_data[i][1]) for i in range(1, len(frames_data))):
            frame_name = frames_data[next(i for i in range(1, len(frames_data)) if int(frames_data[i - 1][1]) > int(frames_data[i][1]))][0]
            return "The frame ("+frame_name+") has a jump in ("+nodes[3]+")."

        return True

    except Exception as e:
        print(f"Error: {e}")
        return False

#Computes the signal's position in the frame
def Signal_Position_inFrame(_position_signal_dbc_, _taille_signal_dbc_):
    
    DIV_ENT = int(_position_signal_dbc_) // 8    
    REST1 = (DIV_ENT + 1) * 8 - int(_position_signal_dbc_)
    REST2 = int(_taille_signal_dbc_) - REST1
    residu = REST2 % 8   
    if residu == 0 :
       k = (REST2 // 8) - 1
    else:
       k = REST2 // 8
    
    if REST2 > 0 :
       _pos_calc_cdc_ = (DIV_ENT - 1 - k) *8 + REST2 - 8 * k - 1
    else:
       _pos_calc_cdc_ = int(_position_signal_dbc_) + int(_taille_signal_dbc_) - 1  
    _Signal_Byte_Position_CDC_ = ( _pos_calc_cdc_ // 8 ) + 1
    _Signal_Bit_Position_CDC_ = _pos_calc_cdc_ % 8

    return _Signal_Byte_Position_CDC_, _Signal_Bit_Position_CDC_

#determines the type of the signal based on the length ,val max,val min,resolution and offset
def Signal_Type(_longueur_, _Signal_phy_min_, _Signal_phy_max_, _Signal_Phy_Resolution_, _Signal_Offset_):
    if int(float(_longueur_))  == 1:
       _type_ = "BOOLEAN"
    elif _Signal_phy_min_ != "Non applicable" or _Signal_phy_max_ != "Non applicable" :
 
        SUP = (float(_Signal_phy_max_) - float(_Signal_Offset_)) // float (_Signal_Phy_Resolution_)
        INF = (float(_Signal_phy_min_) - float(_Signal_Offset_)) // float (_Signal_Phy_Resolution_)
        if (SUP >=0 and  SUP <= 255)  and (INF >=0 and  INF <= 255):
           _type_ = "UINT8"
        elif SUP <= 127 and  INF >= -128:
           _type_ = "SINT8"	
        elif (SUP >=0 and  SUP <= 65535)  and (INF >=0 and  INF <= 65535):	
           _type_ = "UINT16"	   
        elif SUP <= 32767 and  INF >= (-32768):
           _type_ = "SINT16"
        elif (SUP >=0 and  SUP <= 4294967295)  and (INF >=0 and  INF <= 4294967295):			  
           _type_ = "UINT32"		  
        elif SUP <= 2147483647 and  INF >= (-2147483648):
           _type_ = "SINT32"	
    elif int(float(_longueur_))  <= 8:
       _type_ = "UINT8"
    elif int(float(_longueur_))  <= int(16):
       _type_ = "UINT16"	
    elif int(float(_longueur_))  <= int(32):
       _type_ = "UINT32"
    elif int(float(_longueur_))  <= int(64):
       _type_ = "UINT64"	   
    return _type_
