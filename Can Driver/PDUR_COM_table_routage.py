import pandas as pd
import tkinter as tk
import os
from tkinter import filedialog
from lxml import etree
from openpyxl import load_workbook

file_path = os.path.join(os.getcwd(), 'Output.xlsx')

# Function to write data to Excel file
def write_to_Excel(result_data, file_path):
    df = pd.DataFrame(result_data)

    if not os.path.exists(file_path):
        # Create the Excel file with the specified columns
        df.to_excel(file_path, sheet_name='PDUR_COM_table_routage', index=False, header=True)
    else:
        # Load the existing workbook
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book

        if 'PDUR_COM_table_routage' in pd.ExcelFile(file_path).sheet_names:
            sheet = book['PDUR_COM_table_routage']
            # Append the data to the existing sheet
            df.to_excel(writer, sheet_name='PDUR_COM_table_routage', index=False, header=False, startrow=writer.sheets['PDUR_COM_table_routage'].max_row)

        else:
            # Create a new sheet if it doesn't exist
            df.to_excel(writer, sheet_name='PDUR_COM_table_routage', index=False, header=True)

        writer.save()

# Function to extract necessary attributes for the target frame from the .xdm file
def extract_PdurValues(xdm_file, frame_name):
    with open(xdm_file, 'r') as file:
        xdm_content = file.read()

    root = etree.fromstring(xdm_content)
    frame_type=None
    namespace = {'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd','a':'http://www.tresos.de/_projects/DataModel2/08/attribute.xsd'}
    ctr_elements_Tx = root.xpath(".//d:ctr[@name='Com_PduRRoutingTable']/d:lst[@name='PduRRoutingPath']/d:ctr[@name=$name]", namespaces=namespace, name=frame_name)
    ctr_elements_Rx = root.xpath(".//d:ctr[@name='CanIf_PduRRoutingTable']/d:lst[@name='PduRRoutingPath']/d:ctr[@name=$name]", namespaces=namespace, name=frame_name)
    if ctr_elements_Tx and not ctr_elements_Rx:
        
        src_elements = root.xpath(".//d:ctr[@name='Com_PduRRoutingTable']/d:lst[@name='PduRRoutingPath']/d:ctr[@name=$name]/d:ctr[@name='PduRSrcPdu']", namespaces=namespace, name=frame_name)
        dest_elements = root.xpath(".//d:ctr[@name='Com_PduRRoutingTable']/d:lst[@name='PduRRoutingPath']/d:ctr[@name=$name]/d:lst[@name='PduRDestPdu']/d:ctr[@name=$name2]", namespaces=namespace, name=frame_name,name2=frame_name+'_Dest')
        frame_type="Tx"

    elif ctr_elements_Rx and not ctr_elements_Tx:
        src_elements = root.xpath(".//d:ctr[@name='CanIf_PduRRoutingTable']/d:lst[@name='PduRRoutingPath']/d:ctr[@name=$name]/d:ctr[@name='PduRSrcPdu']", namespaces=namespace, name=frame_name)
        dest_elements = root.xpath(".//d:ctr[@name='CanIf_PduRRoutingTable']/d:lst[@name='PduRRoutingPath']/d:ctr[@name=$name]/d:lst[@name='PduRDestPdu']/d:ctr[@name=$name2]", namespaces=namespace, name=frame_name,name2=frame_name+'_Dest')
        frame_type="Rx"
        
    else:
        PduRSrcPdu=PduRSrcBswModuleRef=PduRSrcPduRef=PduRSrcPduUpTxConf=PduRTransmissionConfirmation=PduRDestPduDataProvision=PduRDestBswModuleRef=PduRDestPduRef= None
        return None,PduRSrcPdu, PduRSrcBswModuleRef, PduRSrcPduRef,PduRSrcPduUpTxConf,PduRTransmissionConfirmation,PduRDestPduDataProvision,PduRDestBswModuleRef,PduRDestPduRef

    PduRSrcPdu = src_elements[0].xpath("string(a:a[1]/@value)", namespaces=namespace)
    PduRSrcBswModuleRef = src_elements[0].xpath("string(d:ref[1]/@value)", namespaces=namespace)
    PduRSrcPduRef = src_elements[0].xpath("string(d:ref[2]/@value)", namespaces=namespace)
    PduRSrcPduUpTxConf=src_elements[0].xpath("string(d:var[3]/@value)", namespaces=namespace)
    PduRTransmissionConfirmation=dest_elements[0].xpath("string(d:var[2]/@value)", namespaces=namespace)
    
    PduRDestPduDataProvision=dest_elements[0].xpath("string(d:var[3]/@value)", namespaces=namespace)
    PduRDestBswModuleRef=dest_elements[0].xpath("string(d:ref[1]/@value)", namespaces=namespace)
    PduRDestPduRef=dest_elements[0].xpath("string(d:ref[2]/@value)", namespaces=namespace)
    return frame_type,PduRSrcPdu, PduRSrcBswModuleRef, PduRSrcPduRef,PduRSrcPduUpTxConf,PduRTransmissionConfirmation,PduRDestPduDataProvision,PduRDestBswModuleRef,PduRDestPduRef

def verify_frame(xdm_file_path, frame_name):
    try:
        frame_type,PduRSrcPdu, PduRSrcBswModuleRef, PduRSrcPduRef,PduRSrcPduUpTxConf,PduRTransmissionConfirmation,PduRDestPduDataProvision,PduRDestBswModuleRef,PduRDestPduRef = extract_PdurValues(xdm_file_path, frame_name)
        if frame_type == None and PduRSrcPdu == None and PduRSrcBswModuleRef == None and PduRSrcPduRef == None and PduRSrcPduUpTxConf == None and PduRTransmissionConfirmation == None and PduRDestPduDataProvision == None and PduRDestBswModuleRef == None and PduRDestPduRef == None:
            result_data = {
                'Frame Name': [frame_name],
                'Passed?':["Frame Not Found in PDUR"],
                'Frame Type':[frame_type],
                'PduRSrcPdu':' ',
                'PduRSrcPduUpTxConf':' ',
                'PduRSrcPduRef':' ',
                'PduRSrcBswModuleRef':' ',
                'PduRTransmissionConfirmation':' ',
                'PduRDestPduRef':' ',
                'PduRDestPduDataProvision':' ',
                'PduRDestBswModuleRef':' '
                }
            write_to_Excel(result_data,file_path)
            return False
        else:
            PduRSrcPdutst=PduRSrcBswModuleReftst=PduRSrcPduReftst=PduRSrcPduUpTxConftst=PduRTransmissionConfirmationtst=PduRDestPduDataProvisiontst=PduRDestBswModuleReftst=PduRDestPduReftst=True
            #RX and Tx
            #src
            if(PduRSrcPdu!=frame_name+"_Src"):
                    PduRSrcPdutst=False
            if(PduRSrcPduUpTxConf!="true"):
                    PduRSrcPduUpTxConftst=False
            if(frame_name not in PduRSrcPduRef):
                    PduRSrcPduReftst=False
            #dest
            if(PduRTransmissionConfirmation!="true"):
                    PduRTransmissionConfirmationtst=False
            if(frame_name not in PduRDestPduRef):
                    PduRDestPduReftst=False

            if(frame_type=="Tx"):
                #src specific
                if(PduRSrcBswModuleRef!="/PduR/PduR/BswMod_Com"):
                    PduRSrcBswModuleReftst=False
                
                #dest specific
                
                if(PduRDestPduDataProvision!="PDUR_DIRECT"):
                    PduRDestPduDataProvisiontst=False
                if(PduRDestBswModuleRef!="/PduR/PduR/BswMod_CanIf"):
                    PduRDestBswModuleReftst=False
            
            elif(frame_type=="Rx"):
                #src specific
                if(PduRSrcBswModuleRef!="/PduR/PduR/BswMod_CanIf"):
                    PduRSrcBswModuleReftst=False
                #dest specific
                if(PduRDestBswModuleRef!="/PduR/PduR/BswMod_Com"):
                    PduRDestBswModuleReftst=False
                if(PduRDestPduDataProvision!="PDUR_UPPER"):
                    PduRDestPduDataProvisiontst=False
            print(PduRSrcPdutst, PduRSrcBswModuleReftst, PduRSrcPduReftst, PduRSrcPduUpTxConftst, PduRTransmissionConfirmationtst, PduRDestPduDataProvisiontst, PduRDestBswModuleReftst, PduRDestPduReftst)
            result_data = {
                'Frame Name': [frame_name],
                'Passed?':[" " if PduRSrcPdutst == False or PduRSrcBswModuleReftst == False or PduRSrcPduReftst == False or PduRSrcPduUpTxConftst == False or PduRTransmissionConfirmationtst == False or PduRDestPduDataProvisiontst == False or PduRDestBswModuleReftst == False or PduRDestPduReftst == False else "X"],
                'Frame Type':[frame_type],
                'PduRSrcPdu':[PduRSrcPdu if PduRSrcPdutst else "Error(PduRSrcPdu is not "+frame_name+"_Src"+")"],
                'PduRSrcPduUpTxConf':[PduRSrcPduUpTxConf if PduRSrcPduUpTxConftst else "Error(PduRSrcPduUpTxConf is not of the value 'true'"],
                'PduRSrcPduRef':[PduRSrcPduRef if PduRSrcPduReftst else "Error(PduRSrcPduRef Mismatch)"],
                'PduRSrcBswModuleRef':["Error(PduRSrcBswModuleRef is not '/PduR/PduR/BswMod_Com' for Tx frame )" if PduRSrcBswModuleReftst==False and frame_name=="Tx" else "Error(PduRSrcBswModuleRef is not '/PduR/PduR/BswMod_CanIf' for Rx frame )" if  PduRSrcBswModuleReftst==False and frame_name=="Rx" else PduRSrcBswModuleRef ],
                'PduRTransmissionConfirmation':[PduRTransmissionConfirmation if PduRTransmissionConfirmationtst else "Error(PduRTransmissionConfirmation is not of the value 'true'"],
                'PduRDestPduRef':[PduRDestPduRef if PduRDestPduReftst else "Error(PduRDestPduRef Mismatch)"],
                'PduRDestPduDataProvision':["Error(PduRDestPduDataProvision is not 'PDUR_DIRECT' for Tx frame )" if PduRDestPduDataProvisiontst==False and frame_name=="Tx" else "Error(PduRDestPduDataProvision is not 'PDUR_UPPER' for Rx frame )" if  PduRDestPduDataProvisiontst==False and frame_name=="Rx" else PduRDestPduDataProvision ],
                'PduRDestBswModuleRef':["Error(PduRDestBswModuleRef is not '/PduR/PduR/BswMod_CanIf' for Tx frame )" if PduRDestBswModuleReftst==False and frame_name=="Tx" else "Error(PduRDestBswModuleRef is not '/PduR/PduR/BswMod_Com' for Rx frame )" if  PduRDestBswModuleReftst==False and frame_name=="Rx" else PduRDestBswModuleRef ]
            }
            write_to_Excel(result_data,file_path)

    except Exception as e:
                print(f"Error occurred : {e}")
                return False     


# Clear the Excel file
def clear_excel():
    sheet_name='PDUR_COM_table_routage'
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            sheet.delete_rows(2, sheet.max_row)
        book.save(file_path)
    completion_label.config(text="Output File Cleared", fg="blue")


#select the xdm file from the interface
def browse_xdm():
    xdm_file_path = filedialog.askopenfilename(filetypes=[("XDM files", "*.xdm")])
    if not xdm_file_path:
        return
    xdm_file_entry.delete(0, tk.END)
    xdm_file_entry.insert(tk.END, xdm_file_path)


#execute functionality on button click
def verify_button_click():
    xdm_file_path = xdm_file_entry.get()
    frame_name = frame_entry.get()
    verify_frame(xdm_file_path, frame_name)
    completion_label.config(text="Output Created", fg="green")


# Create the GUI
root = tk.Tk()
root.title("PDUR_COM_table_routage Verification")

frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

xdm_file_label = tk.Label(frame, text="Select PDUR File:")
xdm_file_label.grid(row=1, column=0, padx=5, pady=5)

xdm_file_entry = tk.Entry(frame)
xdm_file_entry.grid(row=1, column=1, padx=5, pady=5)

xdm_file_button = tk.Button(frame, text="Browse", command=browse_xdm)
xdm_file_button.grid(row=1, column=2, padx=5, pady=5)

frame_label = tk.Label(frame, text="Enter Frame Name:")
frame_label.grid(row=2, column=0, padx=5, pady=5)

frame_entry = tk.Entry(frame)
frame_entry.grid(row=2, column=1, padx=5, pady=5)

verify_button = tk.Button(frame, text="Verify", command=verify_button_click)
verify_button.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

clear_excel_button = tk.Button(frame, text="Clear Excel", command=clear_excel)
clear_excel_button.grid(row=6, column=0, columnspan=3, padx=5, pady=5)

completion_label = tk.Label(frame, text="", fg="green")
completion_label.grid(row=7, column=0, columnspan=3, padx=5, pady=5)

root.mainloop()
