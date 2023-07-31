import pandas as pd
import tkinter as tk
import os
from tkinter import filedialog
from lxml import etree
from openpyxl import load_workbook

file_path = os.path.join(os.getcwd(), 'Output.xlsx')

# Define expected headers for cleaning the Excel data
expected_headers = {'FRAMES': ['Radical', 'Activation trame', 'Protocole_M', 'Identifiant_T', 'Taille_Max_T', 'Lmin_T', 'Mode_Transmission_T', 'Nature_Evenement_FR_T', 'Nature_Evenement_GB_T', 'Periode_T', 'UCE Emetteur', 'AEE10r3 Reseau_T']}

# Function to clean the Excel data and keep only the necessary columns
def cleanExcelData(excel_file):
    df = pd.read_excel(excel_file, sheet_name='FRAMES', header=0)
    headers = [col for col in df.columns if col in expected_headers['FRAMES']]
    return df[headers]

# Function to write data to Excel file
def write_to_Excel(result_data, file_path):
    df = pd.DataFrame(result_data)

    if not os.path.exists(file_path):
        # Create the Excel file with the specified columns
        df.to_excel(file_path, sheet_name='CAN_verif_XDM_Messagerie', index=False, header=True)
    else:
        # Load the existing workbook
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book

        if 'CAN_verif_XDM_Messagerie' in pd.ExcelFile(file_path).sheet_names:
            sheet = book['CAN_verif_XDM_Messagerie']
            # Append the data to the existing sheet
            df.to_excel(writer, sheet_name='CAN_verif_XDM_Messagerie', index=False, header=False, startrow=writer.sheets['CAN_verif_XDM_Messagerie'].max_row)

        else:
            # Create a new sheet if it doesn't exist
            df.to_excel(writer, sheet_name='CAN_verif_XDM_Messagerie', index=False, header=True)

        writer.save()


# Function to extract necessary attributes for the target frame from the .xdm file
def extract_CanValues(xdm_file, frame_name):
    with open(xdm_file, 'r') as file:
        xml_content = file.read()

    root = etree.fromstring(xml_content)
    namespace = {'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'}

    ctr_elements = root.xpath(".//d:lst[@name='CanHardwareObject']/d:ctr[contains(@name, $name)]", namespaces=namespace, name=frame_name)
    if ctr_elements:
        CanIdValue = int(ctr_elements[0].xpath("d:var[@name='CanIdValue']/@value", namespaces=namespace)[0])
        CanObjectType = ctr_elements[0].xpath("string(d:var[@name='CanObjectType']/@value)", namespaces=namespace)
        CanIdType = ctr_elements[0].xpath("string(d:var[@name='CanIdType']/@value)", namespaces=namespace)
        CanHandleType = ctr_elements[0].xpath("string(d:var[@name='CanHandleType']/@value)", namespaces=namespace)
        CanControllerRef = ctr_elements[0].xpath("string(d:ref[@name='CanControllerRef']/@value)", namespaces=namespace)
        CanFilterMaskRef = ctr_elements[0].xpath(".//d:lst[@name='CanFilterMaskRef']/d:ref/@value", namespaces=namespace)
    else:
        CanIdValue, CanObjectType, CanIdType, CanHandleType, CanControllerRef,CanFilterMaskRef = None, None, None, None, None,None

    return CanIdValue, CanObjectType, CanIdType, CanHandleType, CanControllerRef,CanFilterMaskRef


# Function to verify the frame attributes from the Excel file with the attributes from the .xdm file
def verify_frame(excel_file_path, xdm_file_path, frame_name):
    CanIdValue, CanObjectType, CanIdType, CanHandleType, CanControllerRef,CanFilterMaskRef = extract_CanValues(xdm_file_path, frame_name)
    if CanIdValue is None and CanObjectType is None and CanIdType is None and CanHandleType is None and CanControllerRef is None and CanFilterMaskRef is None :
        result_data = {
            'Frame Name': [frame_name],
            'Passed?': ["Frame Not Found in Can.xdm File"],
            'CanIdValue': [" "],
            'Identifiant_t': [" "],
            'CanIdValue/Identifiant_t': [" "],
            'CanObjectType': [" "],
            'UCE_Emetteur': [" "],
            'CanObjectType/UCE_Emetteur': [" "],
            'CanIdType': [" "],
            'CanHandleType': [" "],
            'CanControllerRef': [" "],
            'CanFilterMaskRef': [" "],
            'AEE10r3 Reseau_T': [" "],
            'CanControllerRef/AEE10r3 Reseau_T': [" "],
            'CanFilterMaskRef/AEE10r3 Reseau_T': [" "],
        }
        write_to_Excel(result_data,file_path)
        return False
        
        
    frames_data = cleanExcelData(excel_file_path)
    selected_frame = frames_data[frames_data['Radical'] == frame_name]

    if selected_frame.empty:
        result_data = {
            'Frame Name': [frame_name],
            'Passed?': ["Frame Not Found in Messagerie "],
            'CanIdValue': [" "],
            'Identifiant_t': [" "],
            'CanIdValue/Identifiant_t': [" "],
            'CanObjectType': [" "],
            'UCE_Emetteur': [" "],
            'CanObjectType/UCE_Emetteur': [" "],
            'CanIdType': [" "],
            'CanHandleType': [" "],
            'CanControllerRef': [" "],
            'CanFilterMaskRef': [" "],
            'AEE10r3 Reseau_T': [" "],
            'CanControllerRef/AEE10r3 Reseau_T': [" "],
            'CanFilterMaskRef/AEE10r3 Reseau_T': [" "],
        }
        write_to_Excel(result_data,file_path)
        return False

    CanIdValuetst=CanIdTypetst=CanHandleTypetst=CanObjectTypetst=CanControllerReftst=CanFilterMaskReftst=True

    identifiant_t_hex = selected_frame["Identifiant_T"].values[0]
    identifiant_t_decimal = int(identifiant_t_hex, 16)

    if identifiant_t_decimal != CanIdValue:
        CanIdValuetst=False

    if CanIdType != "STANDARD":
        CanIdTypetst=False

    if CanHandleType != "FULL":
        CanHandleTypetst =False
    
    if CanObjectType == "RECEIVE" and not selected_frame["UCE Emetteur"].str.endswith("E_VCU").any():
        pass
    elif CanObjectType == "TRANSMIT" and selected_frame["UCE Emetteur"].str.endswith("E_VCU").any():
        pass
    else:
        CanObjectTypetst=False
    try:

        if "ASPath:/Can/Can/CanConfigSet_0/CAN_1" == CanControllerRef and selected_frame["AEE10r3 Reseau_T"].values[0].startswith("HS1"):
            pass
        elif "ASPath:/Can/Can/CanConfigSet_0/CAN_2" == CanControllerRef and selected_frame["AEE10r3 Reseau_T"].values[0].startswith("HS2"):
            pass
        elif "ASPath:/Can/Can/CanConfigSet_0/CAN_3" == CanControllerRef and selected_frame["AEE10r3 Reseau_T"].values[0].startswith("E_CAN"):
            pass
        else:
            CanControllerReftst=False
        
        if "ASPath:/Can/Can/CanConfigSet_0/CAN_1/AcceptCanIDonly" in CanFilterMaskRef and selected_frame["AEE10r3 Reseau_T"].values[0].startswith("HS1"):
            pass
        elif "ASPath:/Can/Can/CanConfigSet_0/CAN_2/AcceptCanIDonly" in CanFilterMaskRef and selected_frame["AEE10r3 Reseau_T"].values[0].startswith("HS2"):
            pass
        elif "ASPath:/Can/Can/CanConfigSet_0/CAN_3" in CanFilterMaskRef and selected_frame["AEE10r3 Reseau_T"].values[0].startswith("E_CAN"):
            pass
        else:
            CanFilterMaskReftst=False

    except Exception as e:
        print(f"Error occurred : {e}")
        return False   

    result_data = {
        'Frame Name': [frame_name],
        'Passed?':[" " if CanIdValuetst ==False or CanIdTypetst==False or CanHandleTypetst==False or CanObjectTypetst==False or CanControllerReftst==False or CanFilterMaskReftst==False else "X"],
        'CanIdValue':[CanIdValue],
        'Identifiant_t':[identifiant_t_decimal],
        'CanIdValue/Identifiant_t':['Error (ID Mismatch)' if CanIdValuetst==False else " "],
        'CanObjectType': [CanObjectType],
        'UCE_Emetteur': [selected_frame["UCE Emetteur"].values[0]],
        'CanObjectType/UCE_Emetteur': ['Error (CanObjectType Mismatch)' if CanObjectTypetst==False else " "],
        'CanIdType': ['Error (CanIdType is not STANDARD)' if CanIdTypetst==False else CanIdType],
        'CanHandleType': ['Error (CanHandleType is not FULL)' if CanHandleTypetst==False else CanHandleType],
        'CanControllerRef': [CanControllerRef],
        'CanFilterMaskRef': [CanFilterMaskRef],
        'AEE10r3 Reseau_T': [selected_frame["AEE10r3 Reseau_T"].values[0]],
        'CanControllerRef/AEE10r3 Reseau_T': ['Error (CanControllerRef Mismatch)' if CanControllerReftst==False else " "],
        'CanFilterMaskRef/AEE10r3 Reseau_T': ['Error (CanFilterMaskRef Mismatch)' if CanFilterMaskReftst==False else " "],
    }
    write_to_Excel(result_data,file_path)


# Clear the Excel file
def clear_excel():
    file_path = os.path.join(os.getcwd(), 'Output.xlsx')

    # Create an Excel writer object
    excel_writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
    df = pd.DataFrame(columns=['Frame Name',
    'Passed?',
    'CanIdValue',
    'Identifiant_t',
    'CanIdValue/Identifiant_t',
    'CanObjectType',
    'UCE_Emetteur',
    'CanObjectType/UCE_Emetteur',
    'CanIdType',
    'CanHandleType',
    'CanControllerRef',
    'CanFilterMaskRef',
    'AEE10r3 Reseau_T',
    'CanControllerRef/AEE10r3 Reseau_T',
    'CanFilterMaskRef/AEE10r3 Reseau_T',])
    df.to_excel(excel_writer, sheet_name='CAN_verif_XDM_Messagerie', index=False)
    excel_writer.save()


#select the excel file from the interface
def browse_excel():
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not excel_file_path:
        return
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(tk.END, excel_file_path)


#select the xdm file from the interface
def browse_xdm():
    xdm_file_path = filedialog.askopenfilename(filetypes=[("XDM files", "*.xdm")])
    if not xdm_file_path:
        return
    xdm_file_entry.delete(0, tk.END)
    xdm_file_entry.insert(tk.END, xdm_file_path)


#execute functionality on button click
def verify_button_click():
    excel_file_path = excel_file_entry.get()
    xdm_file_path = xdm_file_entry.get()
    frame_name = frame_entry.get()

    verify_frame(excel_file_path, xdm_file_path, frame_name)

# Create the GUI
root = tk.Tk()
root.title("Frame Info XDM/Messagerie Verification")

frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

excel_file_label = tk.Label(frame, text="Select Excel File:")
excel_file_label.grid(row=0, column=0, padx=5, pady=5)

excel_file_entry = tk.Entry(frame)
excel_file_entry.grid(row=0, column=1, padx=5, pady=5)

excel_file_button = tk.Button(frame, text="Browse", command=browse_excel)
excel_file_button.grid(row=0, column=2, padx=5, pady=5)

xdm_file_label = tk.Label(frame, text="Select Can File:")
xdm_file_label.grid(row=1, column=0, padx=5, pady=5)

xdm_file_entry = tk.Entry(frame)
xdm_file_entry.grid(row=1, column=1, padx=5, pady=5)

xdm_file_button = tk.Button(frame, text="Browse", command=browse_xdm)
xdm_file_button.grid(row=1, column=2, padx=5, pady=5)

frame_label = tk.Label(frame, text="Enter Frame Name:")
frame_label.grid(row=2, column=0, padx=5, pady=5)

frame_entry = tk.Entry(frame)
frame_entry.grid(row=2, column=1, padx=5, pady=5)

verify_button = tk.Button(frame, text="Verify", command=verify_button_click)
verify_button.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

clear_excel_button = tk.Button(frame, text="Clear Excel", command=clear_excel)
clear_excel_button.grid(row=6, column=0, columnspan=3, padx=5, pady=5)

root.mainloop()
