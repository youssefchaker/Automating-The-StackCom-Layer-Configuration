import tkinter as tk
from tkinter import filedialog
from lxml import etree
import logging
import pandas as pd
import tkinter as tk
import os
from openpyxl import load_workbook

file_path = os.path.join(os.getcwd(), 'Output.xlsx')

def write_to_Excel(result_data, file_path):
    df = pd.DataFrame(result_data)

    if not os.path.exists(file_path):
        # Create the Excel file with the specified columns
        df.to_excel(file_path, sheet_name='CanIf_index_table_order_verif', index=False, header=True)
    else:
        # Load the existing workbook
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book

        if 'CanIf_index_table_order_verif' in pd.ExcelFile(file_path).sheet_names:
            # Check if the 'Passed?' column already exists in the sheet
            sheet = book['CanIf_index_table_order_verif']
            # Append the data to the existing sheet
            df.to_excel(writer, sheet_name='CanIf_index_table_order_verif', index=False, header=False, startrow=writer.sheets['CanIf_index_table_order_verif'].max_row)

        else:
            # Create a new sheet if it doesn't exist
            df.to_excel(writer, sheet_name='CanIf_index_table_order_verif', index=False, header=True)

        writer.save()

# Clear the Excel file
def clear_excel():
    sheet_name='CanIf_index_table_order_verif'
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            sheet.delete_rows(2, sheet.max_row)
        book.save(file_path)
    completion_label.config(text="Output File Cleared", fg="blue")


#Ordering by CanIfRxPduId
def ordered_by_id_Rx(xdm_file):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()
        root = etree.fromstring(xdm_content)
        ctr_elements = root.xpath(".//d:lst[@name='CanIfRxPduCfg']/d:ctr", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'})
        frames_data = [(ctr.attrib['name'], ctr.xpath("string(d:var[@name='CanIfRxPduId']/@value)", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'})) for ctr in ctr_elements]
        frames_data = [(name, obj_id) for name, obj_id in frames_data if obj_id.strip()]
        first_CanIfRxPduId = int(frames_data[0][1])
        if first_CanIfRxPduId != 0:
            return "The first frame's CanIfRxPduId should be '0', but found '{first_CanIfRxPduId}'"

        CanIfRxPduIds = [int(obj_id) for _, obj_id in frames_data]
        if len(CanIfRxPduIds) != len(set(CanIfRxPduIds)):
            duplicates = [frame_name for frame_name, obj_id in frames_data if CanIfRxPduIds.count(int(obj_id)) > 1]
            errorstring=""
            for frame_name in duplicates:
                errorstring=errorstring+"The frame ("+frame_name+") has a duplicate CanIfRxPduId\n"
            return errorstring

        Last_CanIfRxPduIds = int(frames_data[-1][1])
        total_frames = len(frames_data)
        if Last_CanIfRxPduIds != total_frames - 1:
            return "The last frame's CanIfRxPduId should be '{total_frames - 1}', but found '{Last_CanIfRxPduIds}'"

        if any(int(frames_data[i - 1][1]) > int(frames_data[i][1]) for i in range(1, len(frames_data))):
            frame_name = frames_data[next(i for i in range(1, len(frames_data)) if int(frames_data[i - 1][1]) > int(frames_data[i][1]))][0]
            return "The frame ("+frame_name+") has a jump in CanIfRxPduId"

        return True

    except Exception as e:
        print(f"Error: {e}")
        return False

#Ordering by CanIfTxPduId
def ordered_by_id_Tx(xdm_file):
    try:
        with open(xdm_file, 'r') as file:
            xdm_content = file.read()

        root = etree.fromstring(xdm_content)
        ctr_elements = root.xpath(".//d:lst[@name='CanIfTxPduCfg']/d:ctr", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'})
        frames_data = [(ctr.attrib['name'], ctr.xpath("string(d:var[@name='CanIfTxPduId']/@value)", namespaces={'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'})) for ctr in ctr_elements]
        frames_data = [(name, obj_id) for name, obj_id in frames_data if obj_id.strip()]

        first_CanIfTxPduId = int(frames_data[0][1])
        if first_CanIfTxPduId != 0:
            return "The first frame's CanIfRxPduId should be '0', but found '{first_CanIfTxPduId}'"

        CanIfTxPduIds = [int(obj_id) for _, obj_id in frames_data]
        if len(CanIfTxPduIds) != len(set(CanIfTxPduIds)):
            duplicates = [frame_name for frame_name, obj_id in frames_data if CanIfTxPduIds.count(int(obj_id)) > 1]
            errorstring=""
            for frame_name in duplicates:
                errorstring=errorstring+"The frame ("+frame_name+") has a duplicate CanIfRxPduId\n"
            return errorstring

        Last_CanIfTxPduIds = int(frames_data[-1][1])
        total_frames = len(frames_data)
        if Last_CanIfTxPduIds != total_frames - 1:
            return "The last frame's CanIfRxPduId should be '{total_frames - 1}', but found '{Last_CanIfTxPduIds}'"

        if any(int(frames_data[i - 1][1]) > int(frames_data[i][1]) for i in range(1, len(frames_data))):
            frame_name = frames_data[next(i for i in range(1, len(frames_data)) if int(frames_data[i - 1][1]) > int(frames_data[i][1]))][0]
            return "The frame ("+frame_name+") has a jump in CanIfRxPduId"

        return True

    except Exception as e:
        print(f"Error: {e}")
        return False

def check_order():
    xdm_file_path = xdm_file_entry.get()
    if not xdm_file_path:
        return
    result_data = {
        'Passed?':["X" if ordered_by_id_Rx(xdm_file_path)==ordered_by_id_Tx(xdm_file_path) else " "],
        'Order by CanIfRxPduId':[" " if ordered_by_id_Rx(xdm_file_path)==True else ordered_by_id_Rx(xdm_file_path)],
        'Order by CanIfTxPduId':[" " if ordered_by_id_Tx(xdm_file_path)==True else ordered_by_id_Tx(xdm_file_path)]
     }
    write_to_Excel(result_data,file_path)
    completion_label.config(text="Output Created", fg="green")

    
#open the xdm file
def browse_canif():
    xdm_file_path = filedialog.askopenfilename(filetypes=[("XDM files", "*XDM")])
    if not xdm_file_path:
        return
    xdm_file_entry.delete(0, tk.END)
    xdm_file_entry.insert(tk.END, xdm_file_path)

root = tk.Tk()
root.title("CanIf.xdm File Order Checker")

frame = tk.Frame(root, padx=10, pady=10)
frame.pack()

xml_file_label = tk.Label(frame, text="Select CANIf File:")
xml_file_label.grid(row=0, column=0)

xdm_file_entry = tk.Entry(frame)
xdm_file_entry.grid(row=0, column=1)

xdm_file_button = tk.Button(frame, text="Browse", command=browse_canif)
xdm_file_button.grid(row=0, column=2)

check_receive_transmit_button = tk.Button(frame, text="Check Order", command=check_order)
check_receive_transmit_button.grid(row=1, column=0, columnspan=3, pady=5)

clear_excel_button = tk.Button(frame, text="Clear Excel", command=clear_excel)
clear_excel_button.grid(row=2, column=0, columnspan=3, pady=5)

completion_label = tk.Label(frame, text="", fg="green")
completion_label.grid(row=7, column=0, columnspan=3, padx=5, pady=5)

root.mainloop()
