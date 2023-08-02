import pandas as pd
import tkinter as tk
import os
from tkinter import filedialog
from lxml import etree
from openpyxl import load_workbook

file_path = os.path.join(os.getcwd(), 'Output.xlsx')

# Define expected headers for cleaning the Excel data
expected_headers = {'FRAMES': ['Radical', 'Activation trame', 'Protocole_M', 'Identifiant_T', 'Taille_Max_T', 'Lmin_T', 'Mode_Transmission_T', 'Nature_Evenement_FR_T', 'Nature_Evenement_GB_T', 'Periode_T', 'UCE Emetteur', 'AEE10r3 Reseau_T']}

# Function to clean the Excel data and keep only the necessary columns
def cleanExcelData(excel_file):
    df = pd.read_excel(excel_file, sheet_name='FRAMES', header=0)
    headers = [col for col in df.columns if col in expected_headers['FRAMES']]
    return df[headers]

# Function to write data to Excel file
def write_to_Excel(result_data, file_path):
    df = pd.DataFrame(result_data)

    if not os.path.exists(file_path):
        # Create the Excel file with the specified columns
        df.to_excel(file_path, sheet_name='IF_CANIF_couches_sup', index=False, header=True)
    else:
        # Load the existing workbook
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book

        if 'IF_CANIF_couches_sup' in pd.ExcelFile(file_path).sheet_names:
            sheet = book['IF_CANIF_couches_sup']
            # Append the data to the existing sheet
            df.to_excel(writer, sheet_name='IF_CANIF_couches_sup', index=False, header=False, startrow=writer.sheets['IF_CANIF_couches_sup'].max_row)

        else:
            # Create a new sheet if it doesn't exist
            df.to_excel(writer, sheet_name='IF_CANIF_couches_sup', index=False, header=True)

        writer.save()

# Function to extract necessary attributes for the target frame from the .xdm file
def extract_CanifValues(canif_file_path,can_file_path, frame_name,excel_file_path):
    with open(canif_file_path, 'r') as file:
        canif_content = file.read()
    
    frames_data = cleanExcelData(excel_file_path)

    root_canif = etree.fromstring(canif_content)

    namespace = {'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd','a':'http://www.tresos.de/_projects/DataModel2/08/attribute.xsd' }
    selected_frame = frames_data[frames_data['Radical'] == frame_name]
    if selected_frame.empty:
        result_data = {
                    'Frame Name': [frame_name],
                    'Passed?':["Frame Not Found in Messagerie"],
                    'Frame type':' ',
                    'CanIfPduCanId':' ',
                    'CanIdValue':' ',
                    'CanIfPduCanId/CanIdValue':' ',
                    'CanIfPduCanIdType':' ',
                    'CanIfPduId':' ',
                    'CanObjectId':' ',
                    'CanIfPduId/CanObjectId':' ',
                    'CanIfPduRef':' ',
                    'CanIfPduIdRef':' ',
                    'CanIfPduReadNotifyStatus':' ',
                    'CanIfRxPduReadData':' ',
                    'CanIfRxPduDlc':' ',
                    'Frame Size':' ',
                    'CanIfRxPduDlc/Frame Size':' ',
                    'CanIfRxPduUserRxIndicationName':' ',
                    'CanIfRxPduUserRxIndicationUL':' ',
                    'CanIfTxPduPnFilterPdu':' ',
                    'CanIfTxPduType':' ',
                    'CanIfTxPduUserTxConfirmationName':' ',
                    'CanIfTxPduUserTxConfirmationUL':' '
            }
        write_to_Excel(result_data,file_path)
        CanIfPduCanId= CanIfPduCanIdType=CanIfPduId=CanIfPduRef=CanIfPduIdRef=CanIfPduReadNotifyStatus=CanIfRxPduReadData=CanIfRxPduDlc=CanIfRxPduUserRxIndicationName=CanIfRxPduUserRxIndicationUL=CanIfTxPduPnFilterPdu=CanIfTxPduType=CanIfTxPduUserTxConfirmationName=CanIfTxPduUserTxConfirmationUL= -1
        return CanIfPduCanId, CanIfPduCanIdType,CanIfPduId,CanIfPduRef,CanIfPduIdRef,CanIfPduReadNotifyStatus,CanIfRxPduReadData,CanIfRxPduDlc,CanIfRxPduUserRxIndicationName,CanIfRxPduUserRxIndicationUL,CanIfTxPduPnFilterPdu,CanIfTxPduType ,CanIfTxPduUserTxConfirmationName ,CanIfTxPduUserTxConfirmationUL
    else:
        if(not selected_frame["UCE Emetteur"].str.endswith("E_VCU").any()):
            CanIfTxPduPnFilterPdu = -2
            CanIfTxPduType = -2
            CanIfTxPduUserTxConfirmationName = -2
            CanIfTxPduUserTxConfirmationUL = -2
            ctr_elements = root_canif.xpath(".//d:lst[@name='CanIfRxPduCfg']/d:ctr[contains(@name, $name)]", namespaces=namespace, name=frame_name)
            if ctr_elements:
                CanIfPduCanId = int(ctr_elements[0].xpath("d:var[@name='CanIfRxPduCanId']/@value", namespaces=namespace)[0])
                CanIfPduCanIdType = ctr_elements[0].xpath("string(d:var[@name='CanIfRxPduCanIdType']/@value)", namespaces=namespace)
                CanIfPduId = ctr_elements[0].xpath("d:var[@name='CanIfRxPduId']/@value", namespaces=namespace)[0]
                CanIfPduRef = ctr_elements[0].xpath("string(d:ref[@name='CanIfRxPduRef']/@value)", namespaces=namespace)
                CanIfPduIdRef = ctr_elements[0].xpath("string(d:ref[@name='CanIfRxPduHrhIdRef']/@value)", namespaces=namespace)
                CanIfPduReadNotifyStatus = ctr_elements[0].xpath("d:var[@name='CanIfRxPduReadNotifyStatus']/@value", namespaces=namespace)[0]

                CanIfRxPduReadData = ctr_elements[0].xpath("d:var[@name='CanIfRxPduReadData']/@value", namespaces=namespace)[0]
                CanIfRxPduDlc = ctr_elements[0].xpath("d:var[@name='CanIfRxPduDlc']/@value", namespaces=namespace)[0]
                CanIfRxPduUserRxIndicationName = ctr_elements[0].xpath("string(d:var[@name='CanIfRxPduUserRxIndicationName']/@value)", namespaces=namespace)
                CanIfRxPduUserRxIndicationUL = ctr_elements[0].xpath("string(d:var[@name='CanIfRxPduUserRxIndicationUL']/@value)", namespaces=namespace)
            else:
                CanIfPduCanId= CanIfPduCanIdType=CanIfPduId=CanIfPduRef=CanIfPduIdRef=CanIfPduReadNotifyStatus=CanIfRxPduReadData=CanIfRxPduDlc=CanIfRxPduUserRxIndicationName=CanIfRxPduUserRxIndicationUL= None
            return CanIfPduCanId, CanIfPduCanIdType,CanIfPduId,CanIfPduRef,CanIfPduIdRef,CanIfPduReadNotifyStatus,CanIfRxPduReadData,CanIfRxPduDlc,CanIfRxPduUserRxIndicationName,CanIfRxPduUserRxIndicationUL,CanIfTxPduPnFilterPdu,CanIfTxPduType ,CanIfTxPduUserTxConfirmationName ,CanIfTxPduUserTxConfirmationUL
        else:
            CanIfRxPduReadData = -2
            CanIfRxPduDlc = -2
            CanIfRxPduUserRxIndicationName = -2
            CanIfRxPduUserRxIndicationUL = -2
            ctr_elements = root_canif.xpath(".//d:lst[@name='CanIfTxPduCfg']/d:ctr[contains(@name, $name)]", namespaces=namespace, name=frame_name)
            if ctr_elements:
                CanIfPduCanId = int(ctr_elements[0].xpath("d:var[@name='CanIfTxPduCanId']/@value", namespaces=namespace)[0])
                CanIfPduCanIdType = ctr_elements[0].xpath("string(d:var[@name='CanIfTxPduCanIdType']/@value)", namespaces=namespace)
                CanIfPduId = ctr_elements[0].xpath("d:var[@name='CanIfTxPduId']/@value", namespaces=namespace)[0]
                CanIfPduRef = ctr_elements[0].xpath("string(d:ref[@name='CanIfTxPduRef']/@value)", namespaces=namespace)
                CanIfPduIdRef = ctr_elements[0].xpath("string(d:ref[@name='CanIfTxPduHthIdRef']/@value)", namespaces=namespace)
                CanIfPduReadNotifyStatus = ctr_elements[0].xpath("d:var[@name='CanIfTxPduReadNotifyStatus']/@value", namespaces=namespace)[0]


                CanIfTxPduPnFilterPdu = ctr_elements[0].xpath("d:var[@name='CanIfTxPduPnFilterPdu']/a:a[1]/@value", namespaces=namespace)[0]
                CanIfTxPduType = ctr_elements[0].xpath("string(d:var[@name='CanIfTxPduType']/@value)", namespaces=namespace)
                CanIfTxPduUserTxConfirmationName = ctr_elements[0].xpath("string(d:var[@name='CanIfTxPduUserTxConfirmationName']/@value)", namespaces=namespace)
                CanIfTxPduUserTxConfirmationUL = ctr_elements[0].xpath("string(d:var[@name='CanIfTxPduUserTxConfirmationUL']/@value)", namespaces=namespace)
            else:
                CanIfPduCanId= CanIfPduCanIdType=CanIfPduId=CanIfPduRef=CanIfPduIdRef=CanIfPduReadNotifyStatus=CanIfTxPduPnFilterPdu=CanIfTxPduType=CanIfTxPduUserTxConfirmationName=CanIfTxPduUserTxConfirmationUL= None
            return CanIfPduCanId, CanIfPduCanIdType,CanIfPduId,CanIfPduRef,CanIfPduIdRef,CanIfPduReadNotifyStatus,CanIfRxPduReadData,CanIfRxPduDlc,CanIfRxPduUserRxIndicationName,CanIfRxPduUserRxIndicationUL,CanIfTxPduPnFilterPdu,CanIfTxPduType ,CanIfTxPduUserTxConfirmationName ,CanIfTxPduUserTxConfirmationUL
        
        

# Function to extract necessary attributes for the target frame from the .xdm file
def extract_CanValues(can_file_path, frame_name):

    with open(can_file_path, 'r') as file:
        can_content = file.read()

    root_can = etree.fromstring(can_content)
    namespace = {'d': 'http://www.tresos.de/_projects/DataModel2/06/data.xsd'}

    ctr_elements = root_can.xpath(".//d:lst[@name='CanHardwareObject']/d:ctr[contains(@name, $name)]", namespaces=namespace, name=frame_name)
    if ctr_elements:
        CanIdValue = int(ctr_elements[0].xpath("d:var[@name='CanIdValue']/@value", namespaces=namespace)[0])
        CanObjectId = int(ctr_elements[0].xpath("d:var[@name='CanObjectId']/@value", namespaces=namespace)[0])
    else:
        CanIdValue, CanObjectId = None, None

    return CanIdValue, CanObjectId

def verify_frame(excel_file_path, canif_file_path,can_file_path, frame_name):
    try:
        CanIdValue, CanObjectId=extract_CanValues(can_file_path,frame_name)
        CanIfPduCanId, CanIfPduCanIdType,CanIfPduId,CanIfPduRef,CanIfPduIdRef,CanIfPduReadNotifyStatus,CanIfRxPduReadData,CanIfRxPduDlc,CanIfRxPduUserRxIndicationName,CanIfRxPduUserRxIndicationUL,CanIfTxPduPnFilterPdu,CanIfTxPduType ,CanIfTxPduUserTxConfirmationName ,CanIfTxPduUserTxConfirmationUL = extract_CanifValues(canif_file_path,can_file_path, frame_name,excel_file_path)
        if CanIfPduCanId== -1:
            return False
        elif CanIfPduCanId is None and CanIfPduCanIdType is None and CanIfPduId is None and CanIfPduRef is None and CanIfPduIdRef is None and CanIfPduReadNotifyStatus is None :
            result_data = {
                    'Frame Name': [frame_name],
                    'Passed?':["Frame Not Found in CANIF"],
                    'Frame type':' ',
                    'CanIfPduCanId':' ',
                    'CanIdValue':' ',
                    'CanIfPduCanId/CanIdValue':' ',
                    'CanIfPduCanIdType':' ',
                    'CanIfPduId':' ',
                    'CanObjectId':' ',
                    'CanIfPduId/CanObjectId':' ',
                    'CanIfPduRef':' ',
                    'CanIfPduIdRef':' ',
                    'CanIfPduReadNotifyStatus':' ',
                    'CanIfRxPduReadData':' ',
                    'CanIfRxPduDlc':' ',
                    'Frame Size':' ',
                    'CanIfRxPduDlc/Frame Size':' ',
                    'CanIfRxPduUserRxIndicationName':' ',
                    'CanIfRxPduUserRxIndicationUL':' ',
                    'CanIfTxPduPnFilterPdu':' ',
                    'CanIfTxPduType':' ',
                    'CanIfTxPduUserTxConfirmationName':' ',
                    'CanIfTxPduUserTxConfirmationUL':' '
            }
            write_to_Excel(result_data,file_path)
            return False

        elif (CanIdValue is None and  CanObjectId is None):
            result_data = {
                    'Frame Name': [frame_name],
                    'Passed?':["Frame Not Found in CAN"],
                    'Frame type':' ',
                    'CanIfPduCanId':' ',
                    'CanIdValue':' ',
                    'CanIfPduCanId/CanIdValue':' ',
                    'CanIfPduCanIdType':' ',
                    'CanIfPduId':' ',
                    'CanObjectId':' ',
                    'CanIfPduId/CanObjectId':' ',
                    'CanIfPduRef':' ',
                    'CanIfPduIdRef':' ',
                    'CanIfPduReadNotifyStatus':' ',
                    'CanIfRxPduReadData':' ',
                    'CanIfRxPduDlc':' ',
                    'Frame Size':' ',
                    'CanIfRxPduDlc/Frame Size':' ',
                    'CanIfRxPduUserRxIndicationName':' ',
                    'CanIfRxPduUserRxIndicationUL':' ',
                    'CanIfTxPduPnFilterPdu':' ',
                    'CanIfTxPduType':' ',
                    'CanIfTxPduUserTxConfirmationName':' ',
                    'CanIfTxPduUserTxConfirmationUL':' '
            }
            write_to_Excel(result_data,file_path)
            return False
        else:
            CanIfPduCanIdtst=CanIfPduCanIdTypetst=CanIfPduIdtst=CanIfPduReftst=CanIfPduIdReftst=CanIfPduReadNotifyStatustst=True
            CanIfRxPduReadDatatst=CanIfRxPduDlctst=CanIfRxPduUserRxIndicationNametst=CanIfRxPduUserRxIndicationULtst=CanIfTxPduPnFilterPdutst=CanIfTxPduTypetst=CanIfTxPduUserTxConfirmationNametst=CanIfTxPduUserTxConfirmationULtst=None

            frames_data = cleanExcelData(excel_file_path)
            selected_frame = frames_data[frames_data['Radical'] == frame_name]

            if(CanIfPduCanId!=CanIdValue):
                CanIfPduCanIdtst=False

            if(CanIfPduCanIdType!="STANDARD_CAN"):
                CanIfPduCanIdTypetst=False

            if (CanIfPduId!=CanObjectId):
                CanIfPduIdtst=False

            if(frame_name not in CanIfPduRef):
                CanIfPduReftst=False

            if(frame_name not in CanIfPduIdRef):
                CanIfPduIdReftst=False

            if(CanIfPduReadNotifyStatus!="false"):
                CanIfPduReadNotifyStatustst=False

            if(CanIfTxPduPnFilterPdu == -2 and CanIfTxPduType == -2 and CanIfTxPduUserTxConfirmationName ==-2 and  CanIfTxPduUserTxConfirmationUL== -2):
                
                CanIfRxPduReadDatatst=CanIfRxPduDlctst=CanIfRxPduUserRxIndicationNametst=CanIfRxPduUserRxIndicationULtst=True

                frame_size = selected_frame["Taille_Max_T"].values[0]
                
                if(int(CanIfRxPduDlc)!=int(frame_size)):
                    CanIfRxPduDlctst=False

                if(CanIfRxPduReadData!="false"):
                    CanIfRxPduReadDatatst=False

                
                if(CanIfRxPduUserRxIndicationName!="PduR_RxIndication"):
                    CanIfRxPduUserRxIndicationNametst=False
                
                if(CanIfRxPduUserRxIndicationUL!="PDUR"):
                    CanIfRxPduUserRxIndicationULtst=False

            elif(CanIfRxPduReadData ==-2 and CanIfRxPduDlc ==-2 and CanIfRxPduUserRxIndicationName ==-2 and CanIfRxPduUserRxIndicationUL == -2):

                CanIfTxPduPnFilterPdutst=CanIfTxPduTypetst=CanIfTxPduUserTxConfirmationNametst=CanIfTxPduUserTxConfirmationULtst=True

                if(CanIfTxPduPnFilterPdu!="false"):
                    CanIfTxPduPnFilterPdutst=False

                if(CanIfTxPduType!="STATIC"):
                    CanIfTxPduTypetst=False

                if(CanIfTxPduUserTxConfirmationName!="PduR_TxConfirmation"):
                    CanIfTxPduUserTxConfirmationNametst=False

                if(CanIfTxPduUserTxConfirmationUL!="PDUR"):
                    CanIfTxPduUserTxConfirmationULtst=False
                
            result_data = {
                    'Frame Name': [frame_name],
                    'Passed?':["X" if (CanIfPduCanIdtst==True and CanIfPduCanIdTypetst==True and CanIfPduIdtst==True and CanIfPduReftst==True and CanIfPduIdReftst==True and CanIfPduReadNotifyStatustst==True) and ((CanIfRxPduReadDatatst==True and CanIfRxPduDlctst==True and CanIfRxPduUserRxIndicationNametst==True and CanIfRxPduUserRxIndicationUL==True) or (CanIfTxPduPnFilterPdutst==True and CanIfTxPduTypetst==True and CanIfTxPduUserTxConfirmationNametst==True and CanIfTxPduUserTxConfirmationULtst==True)) else " "],
                    'Frame type':["TRANSMIT" if selected_frame["UCE Emetteur"].str.endswith("E_VCU").any() else "RECEIVE" ],
                    'CanIfPduCanId':[CanIfPduCanId],
                    'CanIdValue':[CanIdValue],
                    'CanIfPduCanId/CanIdValue Errors':["Error(ID Mismatch)" if CanIfPduCanIdtst==False else "None"],
                    'CanIfPduCanIdType':["Error(CanIfPduCanIdType is not STANDARD_CAN)" if CanIfPduCanIdTypetst==False else "STANDARD_CAN" ],
                    'CanIfPduId':[CanIfPduId],
                    'CanObjectId':[CanObjectId],
                    'CanIfPduId/CanObjectId Errors':["Error(Frame Order Mismatch)" if CanIfPduIdtst==False else "None"],
                    'CanIfPduRef':["Error(Frame Name not present in CanIfPduRef)" if CanIfPduReftst==False else CanIfPduRef],
                    'CanIfPduIdRef': ["Error(Frame Name not present in CanIfPduIdRef)" if CanIfPduIdReftst==False else CanIfPduIdRef],
                    'CanIfPduReadNotifyStatus':["Error(CanIfPduReadNotifyStatus is not Deactivated)" if CanIfPduReadNotifyStatustst==False else "Deactivated"],
                    'CanIfRxPduReadData':["Error(CanIfRxPduReadData is not Deactivated)" if CanIfRxPduReadDatatst==False  else "Deactivated" if not CanIfRxPduReadDatatst==None else "---" ],
                    'CanIfRxPduDlc':[CanIfRxPduDlc if not CanIfRxPduDlctst==None else "---"],
                    'Frame Size':[frame_size if not CanIfRxPduDlctst==None else "---"],
                    'CanIfRxPduDlc/Frame Size Errors':["Error(Frame Size Mismatch)" if CanIfRxPduDlctst==False else "None" if not CanIfRxPduDlctst==None else "---"],
                    'CanIfRxPduUserRxIndicationName':["Error(CanIfRxPduUserRxIndicationName is not of the value 'PduR_RxIndication')" if CanIfRxPduUserRxIndicationNametst==False else "PduR_RxIndication" if not CanIfRxPduUserRxIndicationNametst==None else "---"],
                    'CanIfRxPduUserRxIndicationUL':["Error(CanIfRxPduUserRxIndicationUL is not of the value 'PDUR')" if CanIfRxPduUserRxIndicationULtst==False else "PDUR" if not CanIfRxPduUserRxIndicationULtst==None else "---"],
                    'CanIfTxPduPnFilterPdu':["Error(CanIfTxPduPnFilterPdu is not Deactivated)" if CanIfTxPduPnFilterPdutst==False else "Deactivated" if not CanIfTxPduPnFilterPdutst==None else "---"],
                    'CanIfTxPduType':["Error(CanIfTxPduType is not STATIC)" if CanIfTxPduTypetst==False else "STANDARD_CAN" if not CanIfTxPduTypetst==None else "---" ],
                    'CanIfTxPduUserTxConfirmationName':["Error(CanIfTxPduUserTxConfirmationName is not of the value 'PduR_TxConfirmation')" if CanIfTxPduUserTxConfirmationNametst==False else "PDUR" if not CanIfTxPduUserTxConfirmationNametst==None else "---"],
                    'CanIfTxPduUserTxConfirmationUL':["Error(CanIfTxPduUserTxConfirmationUL is not of the value 'PDUR')" if CanIfTxPduUserTxConfirmationULtst==False else "PDUR" if not CanIfTxPduUserTxConfirmationULtst==None else "---"]
            }
            write_to_Excel(result_data,file_path)
            return True
                
    except Exception as e:
                print(f"Error occurred : {e}")
                return False

# Clear the Excel file
def clear_excel():

    # Create an Excel writer object
    excel_writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
    df = pd.DataFrame(columns=[
        'Frame Name',
        'Passed?',
        'Frame type',
        'CanIfPduCanId',
        'CanIdValue',
        'CanIfPduCanId/CanIdValue',
        'CanIfPduCanIdType',
        'CanIfPduId',
        'CanObjectId',
        'CanIfPduId/CanObjectId',
        'CanIfPduRef',
        'CanIfPduIdRef',
        'CanIfPduReadNotifyStatus',
        'CanIfRxPduReadData',
        'CanIfRxPduDlc',
        'Frame Size',
        'CanIfRxPduDlc/Frame Size',
        'CanIfRxPduUserRxIndicationName',
        'CanIfRxPduUserRxIndicationUL',
        'CanIfTxPduPnFilterPdu',
        'CanIfTxPduType',
        'CanIfTxPduUserTxConfirmationName',
        'CanIfTxPduUserTxConfirmationUL'
    ])
    df.to_excel(excel_writer, sheet_name='IF_CANIF_couches_sup', index=False)
    excel_writer.save()
    completion_label.config(text="Output File Cleared", fg="blue")


#select the excel file from the interface
def browse_excel():
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not excel_file_path:
        return
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(tk.END, excel_file_path)


#select the can file from the interface
def browse_canif():
    canif_file_path = filedialog.askopenfilename(filetypes=[("XDM files", "*.xdm")])
    if not canif_file_path:
        return
    canif_file_entry.delete(0, tk.END)
    canif_file_entry.insert(tk.END, canif_file_path)

#select the canif file from the interface
def browse_can():
    can_file_path = filedialog.askopenfilename(filetypes=[("XDM files", "*.xdm")])
    if not can_file_path:
        return
    can_file_entry.delete(0, tk.END)
    can_file_entry.insert(tk.END, can_file_path)


#execute functionality on button click
def verify_button_click():
    excel_file_path = excel_file_entry.get()
    canif_file_path = canif_file_entry.get()
    can_file_path = can_file_entry.get()
    frame_name = frame_entry.get()

    verify_frame(excel_file_path, canif_file_path,can_file_path, frame_name)
    completion_label.config(text="Output Created", fg="green")


# Create the GUI
root = tk.Tk()
root.title("Frame Info CANIF/CAN/Messagerie Verification")

frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

excel_file_label = tk.Label(frame, text="Select Excel File:")
excel_file_label.grid(row=0, column=0, padx=5, pady=5)

excel_file_entry = tk.Entry(frame)
excel_file_entry.grid(row=0, column=1, padx=5, pady=5)

excel_file_button = tk.Button(frame, text="Browse", command=browse_excel)
excel_file_button.grid(row=0, column=2, padx=5, pady=5)

canif_file_label = tk.Label(frame, text="Select Canif File:")
canif_file_label.grid(row=1, column=0, padx=5, pady=5)

canif_file_entry = tk.Entry(frame)
canif_file_entry.grid(row=1, column=1, padx=5, pady=5)

canif_file_button = tk.Button(frame, text="Browse", command=browse_canif)
canif_file_button.grid(row=1, column=2, padx=5, pady=5)

can_file_label = tk.Label(frame, text="Select Can File:")
can_file_label.grid(row=2, column=0, padx=5, pady=5)

can_file_entry = tk.Entry(frame)
can_file_entry.grid(row=2, column=1, padx=5, pady=5)

can_file_button = tk.Button(frame, text="Browse", command=browse_can)
can_file_button.grid(row=2, column=2, padx=5, pady=5)

frame_label = tk.Label(frame, text="Enter Frame Name:")
frame_label.grid(row=3, column=0, padx=5, pady=5)

frame_entry = tk.Entry(frame)
frame_entry.grid(row=3, column=1, padx=5, pady=5)

verify_button = tk.Button(frame, text="Verify", command=verify_button_click)
verify_button.grid(row=4, column=0, columnspan=3, padx=5, pady=5)

clear_excel_button = tk.Button(frame, text="Clear Excel", command=clear_excel)
clear_excel_button.grid(row=6, column=0, columnspan=3, padx=5, pady=5)

completion_label = tk.Label(frame, text="", fg="green")
completion_label.grid(row=7, column=0, columnspan=3, padx=5, pady=5)

root.mainloop()           